import os
import time
import shutil
import sys
import re
from difflib import SequenceMatcher
from datetime import datetime
import pandas as pd
import requests
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


TOOL_DIR = os.path.dirname(os.path.abspath(__file__))
ROOT_DIR = os.path.dirname(TOOL_DIR)

TEMPLATE_FILE = os.path.join(TOOL_DIR, "CN Bulk Upload Template.xlsx")
INPUT_FILE_DEFAULT = os.path.join(TOOL_DIR, "FileName (53) (1).xls")

DAILY_REPORTS_DIR = os.path.join(ROOT_DIR, "DailyReports")
REPORT_REPOSITORY_DIR = os.path.join(ROOT_DIR, "Report Repository")

FIXED_CN_TYPE = "Billed"
FIXED_OFFICE = "GUJARAT"
FIXED_BILLING_OFFICE = "GUJARAT"
CONSIGNOR = "Maruti Suzuki India Limited"
RATE_CHART = "Maruti Rate Chart"
LOAD_TYPE = "MSIL"
FIXED_WEIGHT = 1

NAME_THRESHOLD = 0.80
PLACE_THRESHOLD = 0.85
OEM_THRESHOLD = 0.65

RED_FILL = PatternFill(fill_type="solid", start_color="FF9999", end_color="FF9999")


PLACE_ALIASES = {
    "BANGALORE": "BANGLORE",
    "BENGALURU": "BANGLORE",
    "BANGLORE": "BANGLORE",
    "GURGAON": "GURUGRAM",
    "BOMBAY": "MUMBAI",
    "CALCUTTA": "KOLKATA",
    "NEWDELHI": "DELHI",
    "NEW DELHI": "DELHI",
}


def clean_text(val):
    """Remove extra spaces from text"""
    if pd.isna(val):
        return ""
    return str(val).strip()


def clean_upper(val):
    """Convert text to UPPERCASE and remove extra spaces"""
    return clean_text(val).upper()


def clean_key(val):
    """Convert text to UPPERCASE and keep only letters and numbers"""
    return re.sub(r"[^A-Z0-9]", "", clean_upper(val))


def parse_date(value):
    """Convert date to YYYY-MM-DD format"""
    if pd.isna(value):
        return ""
    dt = pd.to_datetime(value, errors="coerce", dayfirst=False)  
    if pd.isna(dt):
        return ""
    return dt.strftime("%Y-%m-%d")

def normalize_name(name):
    """Clean company names: remove PVT/LTD/LIMITED and special characters"""
    if pd.isna(name):
        return ""

    text = str(name).upper().strip()
    text = re.sub(r"\bPRIVATE\s+LIMITED\b", " PVTLTD ", text)
    text = re.sub(r"\bPVT\.?\s*LTD\.?\b", " PVTLTD ", text)
    text = re.sub(r"\bPRIVATE\b", " PVT ", text)
    text = re.sub(r"\bLIMITED\b", " LTD ", text)
    text = re.sub(r"\bLLP\b", " LLP ", text)
    text = re.sub(r"[^A-Z0-9]", "", text)
    return text


def normalize_place(place):
    """Clean place names: apply aliases and remove special characters"""
    if pd.isna(place):
        return ""

    raw = str(place).upper().strip()
    raw = raw.replace(".", "").replace(",", "")
    raw = " ".join(raw.split())
    compact = re.sub(r"[^A-Z0-9]", "", raw)

    if raw in PLACE_ALIASES:
        return clean_key(PLACE_ALIASES[raw])

    if compact in PLACE_ALIASES:
        return clean_key(PLACE_ALIASES[compact])

    return compact


def normalize_oem(value):
    """Normalize OEM names - handle Maruti variants"""
    text = normalize_name(value)
    text = text.replace("MARUTISUZUKIINDIALIMITED", "MARUTISUZUKI")
    text = text.replace("MARUTISUZUKIINDIALTD", "MARUTISUZUKI")
    text = text.replace("MSIL", "MARUTISUZUKI")
    return text


def get_root_name(name):
    """Get first important word from company name (skip common words)"""
    if not name:
        return ""

    text = str(name).upper().strip().replace(".", " ")
    remove_words = {
        "MOTORS", "AUTOMOBILES", "AUTOMOBILE", "MARUTI", "SUZUKI",
        "CARS", "CAR", "PVT", "PRIVATE", "LIMITED", "LTD", "LLP", "AUTO"
    }
    parts = [p for p in text.split() if p not in remove_words]
    return parts[0] if parts else ""


def similarity(a, b):
    """Calculate string similarity (0.0 to 1.0)"""
    if not a or not b:
        return 0.0
    if a == b:
        return 1.0
    if a in b or b in a:
        return 0.96
    return SequenceMatcher(None, a, b).ratio()


def name_similarity(a, b):
    """Calculate similarity between two company names"""
    return similarity(normalize_name(a), normalize_name(b))


def place_similarity(a, b):
    """Calculate similarity between two place names"""
    return similarity(normalize_place(a), normalize_place(b))


def oem_similarity(a, b):
    """Calculate OEM similarity with special handling for Maruti"""
    a_norm = normalize_oem(a)
    b_norm = normalize_oem(b)

    if not a_norm or not b_norm:
        return 0.0
    if a_norm == b_norm:
        return 1.0
    if a_norm in b_norm or b_norm in a_norm:
        return 0.95
    if "MARUTI" in a_norm and "MARUTI" in b_norm:
        return 0.85
    return similarity(a_norm, b_norm)


def get_papl_data():
    """Fetch polygon/area data from API"""
    url = "http://82.25.108.99:8375/retrieveAllPolygonAreas"
    try:
        response = requests.get(url, timeout=20)
        response.raise_for_status()
        data = response.json()
        return data if isinstance(data, list) else []
    except Exception as e:
        print(f"[API] Polygon API failed: {e}")
        return []


def collect_cn_records(obj, records):
    """Recursively find all CN records in API response"""
    if isinstance(obj, dict):
        if "cnNo" in obj:
            records.append(obj)
        for value in obj.values():
            collect_cn_records(value, records)
    elif isinstance(obj, list):
        for item in obj:
            collect_cn_records(item, records)


def get_existing_cn_data():
    """Get list of CNs already uploaded (to avoid duplicates)"""
    url = "http://82.25.108.99:7045/api/cnCreations/cn-dump-filters"

    existing_cn_numbers = set()
    existing_trip_numbers = set()

    try:
        response = requests.post(url, json={}, timeout=120)
        response.raise_for_status()
        data = response.json()

        records = []
        collect_cn_records(data, records)

        print(f"[API] CN API RECORDS FOUND: {len(records)}")
        
        for item in records:
            cn_no = clean_upper(item.get("cnNo"))
            if cn_no:
                existing_cn_numbers.add(cn_no)
                
                trip_match = re.search(r"(\d+)$", cn_no)
                if trip_match:
                    trip_no = trip_match.group(1)
                    existing_trip_numbers.add(trip_no.upper())

        print(f"[API] Existing CN count: {len(existing_cn_numbers)}")
        print(f"[API] Existing Trip count: {len(existing_trip_numbers)}")
        print(f"[API] Sample CNs: {list(existing_cn_numbers)[:5]}")

        return existing_cn_numbers, existing_trip_numbers

    except Exception as e:
        print(f"[API] CN dump API failed: {e}")
        print("[API] Duplicate check skipped. All input rows will be processed.")
        return set(), set()


def find_best_papl_match(input_consignee, input_city, input_oem, papl_data):
    """
    Find best matching location from polygon data
    Uses 3-factor scoring: Place (30%) + Name (40%) + OEM (30%)
    """
    root_name = get_root_name(input_consignee)
    candidates = []

    for area in papl_data:
        geo = clean_text(area.get("geoName"))
        place = clean_text(area.get("placeName"))
        dealer_oem = clean_text(area.get("dealerOEM", ""))

        if not geo or not place:
            continue

        # Calculate similarity scores
        p_score = place_similarity(input_city, place)        # Location match
        n_score = name_similarity(input_consignee, geo)      # Dealer name match
        o_score = oem_similarity(input_oem, dealer_oem) if dealer_oem else 0.5  # OEM match

        # Skip if OEM doesn't match (if dealer has OEM info)
        if dealer_oem and o_score < OEM_THRESHOLD:
            continue

        # Fast match: same root name + good place match + good OEM
        if root_name and get_root_name(geo) == root_name and p_score >= PLACE_THRESHOLD:
            return {
                "geoName": geo,
                "placeName": place,
                "dealerOEM": dealer_oem,
                "place_score": p_score,
                "name_score": n_score,
                "oem_score": o_score,
                "combined": 1.0,
                "used_original_fallback": False,
                "match_type": "ROOT_PLACE_OEM_MATCH",
            }

        # Calculate weighted score: Place (30%) + Name (40%) + OEM (30%)
        combined = (p_score * 0.30) + (n_score * 0.40) + (o_score * 0.30)

        candidates.append({
            "geoName": geo,
            "placeName": place,
            "dealerOEM": dealer_oem,
            "place_score": p_score,
            "name_score": n_score,
            "oem_score": o_score,
            "combined": combined,
            "used_original_fallback": False,
            "match_type": "FUZZY_OEM",
        })

    # Filter 1: Good place match AND good OEM match
    place_matched = [
        c for c in candidates
        if c["place_score"] >= PLACE_THRESHOLD and c["oem_score"] >= OEM_THRESHOLD
    ]

    if place_matched:
        best = max(place_matched, key=lambda x: (x["name_score"], x["combined"]))
        if best["name_score"] >= NAME_THRESHOLD:
            return best

    # Filter 2: All thresholds satisfied
    if candidates:
        best = max(candidates, key=lambda x: x["combined"])
        if best["place_score"] >= 0.75 and best["name_score"] >= NAME_THRESHOLD:
            return best

    # Fallback: No good match found, use original input
    return {
        "geoName": input_consignee,
        "placeName": input_city,
        "dealerOEM": input_oem,
        "place_score": 0.0,
        "name_score": 0.0,
        "oem_score": 0.0,
        "combined": 0.0,
        "used_original_fallback": True,
        "match_type": "ORIGINAL_INPUT",
    }


def prepare_input_df(input_path):
    """Load and validate Maruti shipment file"""
    df = pd.read_excel(input_path)
    df.columns = [str(c).strip() for c in df.columns]
    df = df.dropna(how="all").reset_index(drop=True)

    required_cols = [
        "Trip No",
        "Invoice Date",
        "Destination City",
        "Dealer Name",
        "Regn No",
        "GR Number",
        "INVOICE NO.",
        "Production Model",
        "TRIP FREIGHT",
        "VIN Number",
        "Engine No",
        "Load No",
        "Exclusive Group",
    ]

    missing = []
    existing_keys = {clean_key(c): c for c in df.columns}

    for col in required_cols:
        if clean_key(col) not in existing_keys:
            missing.append(col)

    if missing:
        raise Exception(f"Missing required columns in Maruti input file: {missing}")

    return df


def get_first_value(row, possible_cols):
    """Get first non-empty value from multiple possible column names"""
    for col in possible_cols:
        if col in row.index:
            val = clean_text(row[col])
            if val:
                return val

    for real_col in row.index:
        real_norm = clean_key(real_col)
        for col in possible_cols:
            if clean_key(col) == real_norm:
                val = clean_text(row[real_col])
                if val:
                    return val

    return ""


def add_split_freight(df):
    """
    Calculate split freight
    If same Regn No + Trip No appears 3 times, divide freight by 3
    """
    trip_col = None
    regn_col = None
    freight_col = None

    for c in df.columns:
        ck = clean_key(c)
        if ck == clean_key("Trip No"):
            trip_col = c
        elif ck == clean_key("Regn No"):
            regn_col = c
        elif ck == clean_key("TRIP FREIGHT"):
            freight_col = c

    df[freight_col] = pd.to_numeric(df[freight_col], errors="coerce").fillna(0)

    # Count how many times each (Regn, Trip) pair appears
    df["_TRIP_REPEAT_COUNT"] = df.groupby([regn_col, trip_col])[trip_col].transform("count")
    
    # Divide total freight by repeat count to get split amount
    df["_SPLIT_FREIGHT"] = df[freight_col] / df["_TRIP_REPEAT_COUNT"]

    return df


def map_mm_material(production_model):
    """Map production model codes to car names"""
    if not production_model:
        return ""
    
    model = clean_upper(production_model).strip()
    model_compact = re.sub(r"\s+", "", model)
    
    mappings = {
        "YNC": "CELERIO",
        "YXA": "BREZZA",
        "YED HB": "SWIFT",
        "YEDHB": "SWIFT",
        "YED NB": "NEW DEZIRE",
        "YEDNB": "NEW DEIRE",
        "YFG": "GRAND VITARA",
        "GRANDVITARA": "GRAND VITARA",
        "YDA": "INVICTO",
        "YY811": "E VITARA",
        "EVITARA": "E VITARA",
        "NEW BALENO": "BALENO",
        "NEWBALENO": "BALENO",
        "BALENO": "BALENO",
        "Y17": "VICTORIS",
        "YE FIX": "VICTORIS",
        "YEFIX": "VICTORIS",
        "VICTORIS": "VICTORIS",
        "NEW ALTO K10": "ALTO K10"
    }
    
    # Try exact match first
    for key, value in mappings.items():
        if model == key:
            return value
    
    # Try match without spaces
    for key, value in mappings.items():
        key_compact = re.sub(r"\s+", "", key)
        if model_compact == key_compact:
            return value
    
    # Try partial match (contains)
    for key, value in mappings.items():
        if key in model or model in key:
            return value
    
    return production_model


def move_file_to_repository(src_file):
    """Move file from DailyReports to Report Repository"""
    try:
        file_name = os.path.basename(src_file)
        
        # Extract date from filename
        date_match = re.search(r"(\d{4}-\d{2}-\d{2})", file_name)
        
        if date_match:
            date_str = date_match.group(1)
            repo_subfolder = os.path.join(REPORT_REPOSITORY_DIR, date_str)
        else:
            repo_subfolder = os.path.join(REPORT_REPOSITORY_DIR, "misc")
        
        os.makedirs(repo_subfolder, exist_ok=True)
        
        dest_path = os.path.join(repo_subfolder, file_name)
        
        # If file already exists, add timestamp
        if os.path.exists(dest_path):
            timestamp = datetime.now().strftime("%H%M%S")
            name, ext = os.path.splitext(file_name)
            new_filename = f"{name}_{timestamp}{ext}"
            dest_path = os.path.join(repo_subfolder, new_filename)
        
        shutil.move(src_file, dest_path)
        print(f"[ARCHIVE] Moved: {file_name} → Report Repository/{os.path.basename(repo_subfolder)}/")
        return dest_path
        
    except Exception as e:
        print(f"[ARCHIVE] Error moving file: {e}")
        return None

def archive_old_daily_files(skip_file=None):
    os.makedirs(DAILY_REPORTS_DIR, exist_ok=True)
    os.makedirs(REPORT_REPOSITORY_DIR, exist_ok=True)

    try:
        existing_files = [
            f for f in os.listdir(DAILY_REPORTS_DIR)
            if f.lower().endswith(".xlsx")
        ]

        if not existing_files:
            print("[ARCHIVE] No files to archive")
            return

        print(f"\n[ARCHIVE] Moving old files to Report Repository...")

        for old_file in existing_files:
            #  NEW FILE KO SKIP KARO
            if skip_file and old_file == skip_file:
                continue

            old_file_path = os.path.join(DAILY_REPORTS_DIR, old_file)

            # date extract karo filename se
            match = re.search(r"(\d{4}-\d{2}-\d{2})", old_file)

            if match:
                date_folder = match.group(1)
            else:
                date_folder = "misc"

            repo_subfolder = os.path.join(REPORT_REPOSITORY_DIR, date_folder)
            os.makedirs(repo_subfolder, exist_ok=True)

            dest_path = os.path.join(repo_subfolder, old_file)

            # agar same naam already hai to rename
            if os.path.exists(dest_path):
                timestamp = datetime.now().strftime("%H%M%S")
                name, ext = os.path.splitext(old_file)
                dest_path = os.path.join(repo_subfolder, f"{name}_{timestamp}{ext}")

            shutil.move(old_file_path, dest_path)

            print(f"[ARCHIVE] {old_file} → {date_folder}/")

        print("[ARCHIVE] Done ")

    except Exception as e:
        print(f"[ARCHIVE ERROR]: {e}")


def process_maruti_cn(input_file, filter_date=None):
    """
    Main function to process Maruti CN generation
    """
    print("=" * 60)
    print("MARUTI BANGLORE CN PROCESS STARTED")
    print("=" * 60)
    print("Input file:", input_file)
    print(f"Filter Date: {filter_date}")

    # Load and prepare input data
    df = prepare_input_df(input_file)
    df = add_split_freight(df)

    # Fetch data from APIs
    papl_data = get_papl_data()
    existing_cn_numbers, existing_trip_numbers = get_existing_cn_data()

    # Filter by date if provided
    if filter_date:
        df["_invoice_date_parsed"] = pd.to_datetime(
            df["Invoice Date"], errors="coerce", dayfirst=True
        )
        filter_dt = pd.to_datetime(filter_date, errors="coerce")
        if pd.notna(filter_dt):
            df = df[df["_invoice_date_parsed"].dt.date == filter_dt.date()]

    if df.empty:
        print("No data found after applying filter.")
        return None

    # Load Excel template
    wb = load_workbook(TEMPLATE_FILE)
    ws = wb.active

    # Clear old data from template
    for row_idx in range(2, ws.max_row + 1):
        for col in range(1, 27):
            ws.cell(row_idx, col).value = None
            ws.cell(row_idx, col).fill = PatternFill(fill_type=None)

    excel_row = 2
    skipped_count = 0
    written_count = 0

    # Process each row from input file
    for _, row in df.iterrows():
        trip_no = get_first_value(row, ["Trip No"])
        cn_no = trip_no
        cn_key = clean_upper(cn_no)
        trip_key = clean_upper(trip_no)

        # Skip if this CN/Trip already uploaded
        if cn_key in existing_cn_numbers or trip_key in existing_trip_numbers:
            skipped_count += 1
            print(f"[DUPLICATE] Skipping already uploaded CN: {cn_no} | Trip: {trip_no}")
            continue

        # Extract all required fields from row
        invoice_date = parse_date(get_first_value(row, ["Invoice Date"]))
        destination_city = get_first_value(row, ["Destination City"])
        dealer_name = get_first_value(row, ["Dealer Name"])
        regn_no = get_first_value(row, ["Regn No"])
        gr_number = get_first_value(row, ["GR Number"])
        invoice_no = get_first_value(row, ["INVOICE NO.", "Invoice No"])
        production_model_raw = get_first_value(row, ["Production Model"])
        production_model = map_mm_material(production_model_raw)
        vin_number = get_first_value(row, ["VIN Number"])
        engine_no = get_first_value(row, ["Engine No"])
        load_no = get_first_value(row, ["Load No"])
        exclusive_group = get_first_value(row, ["Exclusive Group"])

        # Get split freight values
        split_freight = float(row["_SPLIT_FREIGHT"]) if pd.notna(row["_SPLIT_FREIGHT"]) else 0
        repeat_count = int(row["_TRIP_REPEAT_COUNT"]) if pd.notna(row["_TRIP_REPEAT_COUNT"]) else 1

        # Find best matching location from polygon data
        papl_match = find_best_papl_match(
            input_consignee=dealer_name,
            input_city=destination_city,
            input_oem=CONSIGNOR,
            papl_data=papl_data,
        )

        final_consignee = papl_match["geoName"]
        final_destination = papl_match["placeName"]
        oem_score = papl_match.get("oem_score", 0.0)
        used_fallback = papl_match.get("used_original_fallback", False)
        match_type = papl_match.get("match_type")

        # Build route
        route = f"MEHSANA-{final_destination}-{CONSIGNOR}"

        # Print processing details
        print("-" * 60)
        print(f"Excel Row {excel_row}")
        print("Trip No:", trip_no)
        print("Regn No:", regn_no)
        print("Repeat Count:", repeat_count)
        print("Original Trip Freight:", get_first_value(row, ["TRIP FREIGHT"]))
        print("Final Rate/Freight:", split_freight)
        print("Input Dealer:", dealer_name)
        print("Input City:", destination_city)
        print("Production Model (Raw):", production_model_raw)
        print("Production Model (Mapped):", production_model)
        print("Matched Destination:", final_destination)
        print("Matched Consignee:", final_consignee)
        print("OEM Score:", f"{oem_score:.2f}")
        print("Match Type:", match_type)
        print("Used Fallback:", used_fallback)

        # Write data to Excel cells
        ws.cell(excel_row, 1).value = cn_no
        ws.cell(excel_row, 2).value = FIXED_CN_TYPE
        ws.cell(excel_row, 3).value = invoice_date
        ws.cell(excel_row, 4).value = FIXED_OFFICE
        ws.cell(excel_row, 5).value = FIXED_BILLING_OFFICE
        ws.cell(excel_row, 6).value = CONSIGNOR
        ws.cell(excel_row, 7).value = route
        ws.cell(excel_row, 8).value = route
        ws.cell(excel_row, 9).value = final_consignee
        ws.cell(excel_row, 10).value = regn_no
        ws.cell(excel_row, 11).value = RATE_CHART
        ws.cell(excel_row, 12).value = LOAD_TYPE
        ws.cell(excel_row, 13).value = gr_number
        ws.cell(excel_row, 14).value = invoice_no
        ws.cell(excel_row, 15).value = invoice_date
        ws.cell(excel_row, 16).value = production_model
        ws.cell(excel_row, 17).value = FIXED_WEIGHT
        ws.cell(excel_row, 18).value = split_freight
        ws.cell(excel_row, 19).value = split_freight
        ws.cell(excel_row, 20).value = ""
        ws.cell(excel_row, 21).value = vin_number
        ws.cell(excel_row, 22).value = engine_no
        ws.cell(excel_row, 23).value = load_no
        ws.cell(excel_row, 24).value = exclusive_group
        ws.cell(excel_row, 25).value = ""
        ws.cell(excel_row, 26).value = ""

        # Highlight in red if fallback match was used
        if used_fallback:
            ws.cell(excel_row, 9).fill = RED_FILL

        excel_row += 1
        written_count += 1

    # Create output filename with timestamp
    run_datetime = datetime.now().strftime("%Y-%m-%d_%H%M%S")
    output_file = f"MARUTI_GUJARAT_CN_{run_datetime}.xlsx"

    # Save Excel file
    wb.save(output_file)

    archive_old_daily_files(skip_file=output_file)

    os.makedirs(DAILY_REPORTS_DIR, exist_ok=True)
    daily_reports_path = os.path.join(DAILY_REPORTS_DIR, output_file)

    shutil.move(output_file, daily_reports_path)

    print(f"\n[PROCESS] Saved to DailyReports: {daily_reports_path}")

    final_output_path = daily_reports_path

    # Print final summary
    print("=" * 60)
    print(f"Saved to DailyReports: {final_output_path}")
    print(f"Rows written: {written_count}")
    print(f"Rows skipped (CN already exists in API): {skipped_count}")
    print("=" * 60)

    return final_output_path


if __name__ == "__main__":
    # Get today's date automatically
    today_date = pd.Timestamp.today().strftime("%Y-%m-%d")
    
    # Get input file from command line or use default
    input_file = sys.argv[1] if len(sys.argv) > 1 else INPUT_FILE_DEFAULT
    
    # Get filter date from command line or use today's date
    filter_date = sys.argv[2] if len(sys.argv) > 2 else today_date
    
    # Print startup information
    print(f"\n{'='*60}")
    print(f"[AUTO-DATE] System Date: {today_date}")
    print(f"[AUTO-DATE] Processing Date: {filter_date}")
    print(f"[AUTO-DATE] Output File: Maruti_CN_Export_{filter_date}_HHMMSS.xlsx")
    print(f"[AUTO-DATE] Location: DailyReports/")
    print(f"[AUTO-DATE] Old Files Archive: Report Repository/{filter_date}/")
    print(f"{'='*60}\n")
    
    # Run the main process
    result = process_maruti_cn(input_file, filter_date=filter_date)
    
    # Print final result
    if result:
        print(f"\n SUCCESS!")
        print(f"Output file saved: {result}")
    else:
        print(f"\n FAILED!")
        print(f"No output generated")