import os
import sys
import re
import time
import shutil
from difflib import SequenceMatcher
from datetime import datetime
import pandas as pd
import requests
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

TOOL_DIR = os.path.dirname(os.path.abspath(__file__))
ROOT_DIR = os.path.dirname(TOOL_DIR)

TEMPLATE_FILE = os.path.join(TOOL_DIR, "CN Bulk Upload Template.xlsx")
INPUT_FILE_DEFAULT = os.path.join(TOOL_DIR, "TATA 1 MONTH SHIPMENT DOWNLOAD.xlsx")
FREIGHT_FILE = os.path.join(TOOL_DIR, "New Freight List From 19.01.26 2026.xlsx")

DAILY_REPORTS_DIR = os.path.join(ROOT_DIR, "DailyReports")
REPORT_REPOSITORY_DIR = os.path.join(ROOT_DIR, "Report Repository")

FIXED_CN_TYPE = "Billed"
FIXED_OFFICE = "Sanand"
FIXED_BILLING_OFFICE = "Sanand"
FIXED_WEIGHT = 1

NON_EV_CONSIGNOR = "TATA MOTORS PV LTD"
EV_CONSIGNOR = "Tata Passenger Electric Mobility Limited"

NON_EV_RATE_CHART = "Tata PV Rate Chart"
EV_RATE_CHART = "Tata EV Rate Chart"

NON_EV_LOAD_TYPE = "TML PV Fixed"
EV_LOAD_TYPE = "TML EV Fixed"

NAME_THRESHOLD = 0.80
PLACE_THRESHOLD = 0.85
OEM_THRESHOLD = 0.70  
FREIGHT_SIM_THRESHOLD = 0.75

RED_FILL = PatternFill(fill_type="solid", start_color="FF9999", end_color="FF9999")

PLACE_ALIASES = {
    "BANGALORE": "BENGALURU",
    "GURGAON": "GURUGRAM",
    "BOMBAY": "MUMBAI",
    "CALCUTTA": "KOLKATA",
    "TRIVANDRUM": "THIRUVANANTHAPURAM",
    "ANANTHAPUR": "ANANTAPUR",
    "HARDWAR": "HARIDWAR",
    "HISSAR": "HISAR",
    "AMARAVATI": "AMRAVATI",
    "VADODARA": "BARODA",
    "BARODA": "BARODA",
    "BADODA": "BARODA",
    "BRODA": "BARODA",
    "NEWDELHI": "DELHI",
    "NEW DELHI": "DELHI",
}

TATA_OVERRIDES = {
    # ("ABC MOTORS PRIVATE LIMITED", "BRODA"): ("ABC MOTORS PVT LTD", "BARODA"),
}

def clean_text(val):
    if pd.isna(val):
        return ""
    return str(val).strip()
def clean_upper(val):
    return clean_text(val).upper()
def clean_key(val):
    return re.sub(r"[^A-Z0-9]", "", clean_upper(val))
def parse_date(value):
    if pd.isna(value):
        return ""
    dt = pd.to_datetime(value, errors="coerce", dayfirst=True)
    if pd.isna(dt):
        return ""
    return dt.strftime("%Y-%m-%d")
def normalize_name(name):
    if pd.isna(name):
        return ""
    text = str(name).upper().strip()
    text = re.sub(r"\bPRIVATE\s+LIMITED\b", " PVTLTD ", text)
    text = re.sub(r"\bPVT\.?\s*LTD\.?\b", " PVTLTD ", text)
    text = re.sub(r"\bPVT\s+LIMITED\b", " PVTLTD ", text)
    text = re.sub(r"\bPRIVATE\b", " PVT ", text)
    text = re.sub(r"\bLIMITED\b", " LTD ", text)
    text = re.sub(r"\bLLP\b", " LLP ", text)
    text = re.sub(r"[^A-Z0-9]", "", text)
    return text
def normalize_place(place):
    if pd.isna(place):
        return ""
    raw = str(place).upper().strip()
    raw = raw.replace(".", "").replace(",", "")
    raw = " ".join(raw.split())
    compact = re.sub(r"[^A-Z0-9]", "", raw)
    if raw in PLACE_ALIASES:
        return clean_key(PLACE_ALIASES[raw])
    if compact in PLACE_ALIASES:
        return clean_key(PLACE_ALIASES[compact])
    return compact
def similarity(a, b):
    if not a or not b:
        return 0.0
    if a == b:
        return 1.0
    if a in b or b in a:
        return 0.96
    return SequenceMatcher(None, a, b).ratio()
def name_similarity(a, b):
    return similarity(normalize_name(a), normalize_name(b))
def place_similarity(a, b):
    return similarity(normalize_place(a), normalize_place(b))

def get_root_name(name):
    if not name:
        return ""
    text = str(name).upper().strip().replace(".", " ")
    remove_words = {
        "MOTORS", "AUTOMOBILES", "AUTOMOBILE", "TATA",
        "CARS", "CAR", "PVT", "PRIVATE", "LIMITED",
        "LLP", "LTD", "PVTLTD", "MOTOR", "AUTO"
    }
    parts = [p for p in text.split() if p not in remove_words]
    return parts[0] if parts else ""

def get_papl_data():
    url = "http://82.25.108.99:8375/retrieveAllPolygonAreas"
    try:
        response = requests.get(url, timeout=10)
        if response.status_code == 200:
            data = response.json()
            return data if isinstance(data, list) else []
        return []
    except Exception as e:
        print(f"Polygon API failed: {e}")
        return []

def extract_shipment_from_cn(cn_no):
    cn = clean_text(cn_no).upper()
    for prefix in ["SAN/EV/", "SAN-", "PUNE/EV/", "PUNE-", "RDR/EV/", "RDR-"]:
        if cn.startswith(prefix):
            return cn.replace(prefix, "").strip()
    match = re.search(r"(\d+)$", cn)
    if match:
        return match.group(1)
    return ""

def collect_cn_records(obj, records):
    if isinstance(obj, dict):
        if "cnNo" in obj:
            records.append(obj)
        for value in obj.values():
            collect_cn_records(value, records)
    elif isinstance(obj, list):
        for item in obj:
            collect_cn_records(item, records)

def get_existing_cn_data():
    url = "http://82.25.108.99:7045/api/cnCreations/cn-dump-filters"
    existing_cn_numbers = set()
    existing_shipments = set()
    try:
        response = requests.post(url, json={},timeout=120)
        print("CN API STATUS:",response.status_code)
        print("CN API RAW START:",response.text[:1000])
        response.raise_for_status()
        data = response.json()
        records = []
        collect_cn_records(data, records)
        print("CN API RECORDS FOUND:", len(records))
        for item in records:
            cn_no = clean_text(item.get("cnNo")).upper()
            if cn_no:
                existing_cn_numbers.add(cn_no)
                shipment = extract_shipment_from_cn(cn_no)
                if shipment:
                    existing_shipments.add(shipment.upper())
        print(f"Existing CN count from API: {len(existing_cn_numbers)}")
        print(f"Existing Shipment count from API: {len(existing_shipments)}")
        print("API sample CN:", list(existing_cn_numbers)[:10])
        return existing_cn_numbers, existing_shipments
    except Exception as e:
        print(f"CN dump API failed: {e}")
        print("Duplicate CN check skipped. All input rows will be processed.")
        return set(), set()

# --- Fuzzy Consignee Implementation (Kia/Hyundai logic) ---
def find_override_match(input_consignee, input_city, input_oem):
    consignee_norm = normalize_name(input_consignee)
    city_norm = normalize_place(input_city)
    for (raw_name, raw_city), (papl_name, papl_city) in TATA_OVERRIDES.items():
        if consignee_norm == normalize_name(raw_name) and city_norm == normalize_place(raw_city):
            return {
                "geoName": papl_name,
                "placeName": papl_city,
                "dealerOEM": input_oem,
                "oem_score": 1.0,
                "place_score": 1.0,
                "name_score": 1.0,
                "combined": 1.0,
                "used_original_fallback": False,
                "match_type": "OVERRIDE",
            }
    return None
def find_root_destination_match(input_consignee, input_city, input_oem, papl_data):
    root_input = get_root_name(input_consignee)
    city_norm_input = normalize_place(input_city)
    if not root_input or not city_norm_input:
        return None
    for area in papl_data:
        geo = str(area.get("geoName", ""))
        place = str(area.get("placeName", ""))
        dealer_oem = str(area.get("dealerOEM", ""))
        if not geo or not place:
            continue
        root_api = get_root_name(geo)
        city_norm_api = normalize_place(place)
        oem_score = name_similarity(input_oem, dealer_oem)
        if (root_api and root_input == root_api and city_norm_input == city_norm_api and oem_score >= OEM_THRESHOLD):
            return {
                "geoName": geo,
                "placeName": place,
                "dealerOEM": dealer_oem,
                "place_score": 1.0,
                "name_score": 1.0,
                "oem_score": oem_score,
                "combined": 1.0,
                "used_original_fallback": False,
                "match_type": "ROOT_NAME_PLACE_OEM",
            }
    return None
def find_best_papl_match(input_consignee, input_city, input_oem, papl_data):
    override = find_override_match(input_consignee, input_city, input_oem)
    if override:
        return override
    root_match = find_root_destination_match(input_consignee, input_city, input_oem, papl_data)
    if root_match:
        return root_match
    candidates = []
    for area in papl_data:
        geo = str(area.get("geoName", ""))
        place = str(area.get("placeName", ""))
        dealer_oem = str(area.get("dealerOEM", ""))
        if not geo or not place:
            continue
        p_score = place_similarity(input_city, place)
        n_score = name_similarity(input_consignee, geo)
        o_score = name_similarity(input_oem, dealer_oem)
        combined = (p_score * 0.30) + (n_score * 0.40) + (o_score * 0.30)
        candidates.append({
            "geoName": geo,
            "placeName": place,
            "dealerOEM": dealer_oem,
            "place_score": p_score,
            "name_score": n_score,
            "oem_score": o_score,
            "combined": combined,
            "used_original_fallback": False,
            "match_type": "FUZZY",
        })
    place_matched = [
        c for c in candidates 
        if c["place_score"] >= PLACE_THRESHOLD and c["oem_score"] >= OEM_THRESHOLD
    ]
    if place_matched:
        best = max(place_matched, key=lambda x: (x["name_score"], x["combined"]))
        if best["name_score"] >= NAME_THRESHOLD:
            return best
    if candidates:
        best = max(candidates, key=lambda x: x["combined"])
        if (best["name_score"] >= NAME_THRESHOLD and 
            best["place_score"] >= 0.75 and
            best["oem_score"] >= 0.70):
            return best
    return {
        "geoName": input_consignee,
        "placeName": input_city,
        "dealerOEM": input_oem,
        "place_score": 0.0,
        "name_score": 0.0,
        "oem_score": 0.0,
        "combined": 0.0,
        "used_original_fallback": True,
        "match_type": "ORIGINAL_INPUT",
    }

def is_ev_vehicle(description):
    d = clean_upper(description)
    ev_keywords = [
        ".EV",
        " EV ",
        "NEXON.EV",
        "TIAGO EV",
        "PUNCH.EV",
        "HARRIER.EV",
        "CURVV.EV",
        "XPRES-T EV",
        "TIGOR",
    ]
    return any(k in d for k in ev_keywords)
def map_mm_material(description):
    d = clean_upper(description)
    if "NEXON.EV" in d:
        return "NEXON EV"
    if "NEXON" in d:
        return "NEXON"
    if "TIAGO EV" in d:
        return "TIAGO EV"
    if "TIAGO" in d:
        return "TIAGO"
    if "SIERRA" in d:
        return "SIERRA"
    if "XPRES-T CNG" in d or "XPRES-T " in d:
        return "TIGOR"
    if "HARRIER.EV" in d:
        return "HARRIER EV"
    if "HARRIER" in d:
        return "HARRIER"
    if "PUNCH.EV" in d:
        return "PUNCH EV"
    if "PUNCH" in d:
        return "PUNCH"
    if "SAFARI" in d:
        return "SAFARI"
    if "ALTROZ" in d:
        return "ALTROZ"
    if "CURVV.EV" in d:
        return "CURVV EV"
    if "CURVV" in d:
        return "CURVV"
    if "TIGOR CNG" in d:
        return "TIGOR"
    if "TIGOR.EV" in d:
        return "TIGOR EV"
    if  "XPRES-T EV" in d:
         return "TIGOR" 
    return clean_text(description)
def get_freight_bucket(mm_material):
    material = clean_upper(mm_material)
    if material in ["TIAGO", "TIAGO EV"]:
        return "Tiago"
    if material in [
        "NEXON", "NEXON EV", "TIGOR",
        "PUNCH", "PUNCH EV", "ALTROZ",
        "CURVV", "CURVV EV",
        "HARRIER", "HARRIER EV",
        "SAFARI"
    ]:
        return "Nexon/ Tiagor"
    if material == "SIERRA":
        return "Sierra"
    return None

def prepare_input_df(input_path):
    df = pd.read_excel(input_path)
    df.columns = [str(c).strip() for c in df.columns]
    df = df.dropna(how="all").reset_index(drop=True)
    required_cols = [
        "Shipment No.",
        "March Date",
        "Dealer Code",
        "Dealer Location",
        "Trailer No.",
        "Invoice No.",
        "Invoice Date",
        "Chassis No.",
        "Description",
    ]
    missing = [c for c in required_cols if c not in df.columns]
    if missing:
        raise Exception(f"Missing required columns in TATA input file: {missing}")
    return df
def prepare_freight_df(master_path):
    xls = pd.ExcelFile(master_path)
    sheet_name = xls.sheet_names[0]
    df = pd.read_excel(master_path, sheet_name=sheet_name)
    df.columns = [str(c).strip() for c in df.columns]
    df = df.dropna(how="all").reset_index(drop=True)
    required_cols = ["Destination"]
    missing = [c for c in required_cols if c not in df.columns]
    if missing:
        raise Exception(f"Missing required columns in freight file: {missing}")
    df["DEST_KEY"] = df["Destination"].apply(clean_key)
    return df

# ⭐️⭐️⭐️ Fuzzy Freight Location Matching ⭐️⭐️⭐️
def get_rate_from_master(destination, mm_material, freight_df):
    dest_key = clean_key(destination)
    bucket_col = get_freight_bucket(mm_material)
    if not bucket_col:
        print(f"Rate bucket not found for material: {mm_material}")
        return 1
    if bucket_col not in freight_df.columns:
        print(f"Rate column not found in freight master: {bucket_col}")
        return 1
    # 1. Try exact match
    match = freight_df[freight_df["DEST_KEY"] == dest_key]
    if not match.empty:
        rate = pd.to_numeric(match.iloc[0][bucket_col], errors="coerce")
        if pd.notna(rate):
            return float(rate)
    # 2. Fuzzy match if destination not found
    max_sim = 0
    rate_val = 1
    for idx, row in freight_df.iterrows():
        sim = place_similarity(destination, row["Destination"])
        if sim > max_sim and sim >= FREIGHT_SIM_THRESHOLD:
            try:
                candidate_rate = float(row[bucket_col])
                if not pd.isna(candidate_rate):
                    max_sim = sim
                    rate_val = candidate_rate
            except:
                continue
    if max_sim >= FREIGHT_SIM_THRESHOLD:
        print(f"Fuzzy rate match for dest: {destination}, sim: {max_sim:.2f}, rate: {rate_val}")
        return rate_val
    print(f"Destination not found in freight master (even fuzzy): {destination}")
    return 1

def archive_old_daily_files():
    os.makedirs(DAILY_REPORTS_DIR, exist_ok=True)
    os.makedirs(REPORT_REPOSITORY_DIR, exist_ok=True)
    today_folder = datetime.now().strftime("%Y-%m-%d")
    repo_subfolder = os.path.join(REPORT_REPOSITORY_DIR, today_folder)
    os.makedirs(repo_subfolder, exist_ok=True)
    existing_files = [
        f for f in os.listdir(DAILY_REPORTS_DIR)
        if f.lower().endswith(".xlsx")
    ]
    for old_file in existing_files:
        old_file_path = os.path.join(DAILY_REPORTS_DIR, old_file)
        file_name_wo_ext, ext = os.path.splitext(old_file)
        timestamp = datetime.now().strftime("%H%M%S")
        dest_path = os.path.join(repo_subfolder, old_file)
        if os.path.exists(dest_path):
            dest_path = os.path.join(
                repo_subfolder,
                f"{file_name_wo_ext}_{timestamp}{ext}"
            )
        shutil.move(old_file_path, dest_path)

def process_tata_cn(input_file, filter_date=None):
    print("=" * 60)
    print("TATA SANAND CN PROCESS STARTED")
    print("=" * 60)
    print("Input file:", input_file)
    print("Freight file:", FREIGHT_FILE)
    df = prepare_input_df(input_file)
    freight_df = prepare_freight_df(FREIGHT_FILE)
    papl_data = get_papl_data()
    existing_cn_numbers, existing_shipments = get_existing_cn_data()
    if filter_date:
        df["parsed_march_date"] = pd.to_datetime(
            df["March Date"], errors="coerce", dayfirst=True
        )
        filter_dt = pd.to_datetime(filter_date, errors="coerce")
        if pd.notna(filter_dt):
            df = df[df["parsed_march_date"].dt.date == filter_dt.date()]
    if df.empty:
        print("No data found after applying filter.")
        return None
    wb = load_workbook(TEMPLATE_FILE)
    ws = wb.active
    for row_idx in range(2, ws.max_row + 1):
        for col in range(1, 27):
            ws.cell(row_idx, col).value = None
            ws.cell(row_idx, col).fill = PatternFill(fill_type=None)
    excel_row = 2
    skipped_count = 0
    written_count = 0
    for _, row in df.iterrows():
        shipment_no = clean_text(row.get("Shipment No."))
        shipment_key = clean_upper(shipment_no)
        cn_date = parse_date(row.get("March Date"))
        dealer_code = clean_text(row.get("Dealer Code"))
        dealer_name = clean_text(row.get("Dealer Name")) if "Dealer Name" in df.columns else ""
        destination_input = clean_text(row.get("Dealer Location"))
        trailer_no = clean_text(row.get("Trailer No."))
        invoice_no = clean_text(row.get("Invoice No."))
        invoice_date = parse_date(row.get("Invoice Date"))
        chassis_no = clean_text(row.get("Chassis No."))
        description = clean_text(row.get("Description"))
        mm_material = map_mm_material(description)
        ev_flag = is_ev_vehicle(description)
        if ev_flag:
            cn_no = f"SAN/EV/{shipment_no}"
            consignor = EV_CONSIGNOR
            rate_chart = EV_RATE_CHART
            load_type = EV_LOAD_TYPE
        else:
            cn_no = f"SAN-{shipment_no}"
            consignor = NON_EV_CONSIGNOR
            rate_chart = NON_EV_RATE_CHART
            load_type = NON_EV_LOAD_TYPE
        cn_key = clean_upper(cn_no)
        if cn_key in existing_cn_numbers or shipment_key in existing_shipments:
            skipped_count += 1
            print(f"Skipping already uploaded CN: {cn_no} | Shipment: {shipment_no}")
            continue
        input_consignee = dealer_name if dealer_name else dealer_code
        papl_match = find_best_papl_match(input_consignee, destination_input, consignor, papl_data)
        final_consignee = papl_match["geoName"]
        final_destination = papl_match["placeName"]
        api_oem = papl_match.get("dealerOEM", "")
        oem_score = papl_match.get("oem_score", 0.0)
        p_score = papl_match.get("place_score", 0.0)
        n_score = papl_match.get("name_score", 0.0)
        combined = papl_match.get("combined", 0.0)
        used_fallback = papl_match.get("used_original_fallback", False)
        match_type = papl_match.get("match_type")
        route = f"Sanand-{final_destination}-{NON_EV_CONSIGNOR}"
        rate = get_rate_from_master(final_destination, mm_material, freight_df)
        freight = rate
        print("-" * 60)
        print(f"Excel Row {excel_row}")
        print("Shipment No.:", shipment_no)
        print("CN No:", cn_no)
        print("Input Dealer Code:", dealer_code)
        print("Input Dealer Name:", dealer_name)
        print("Input Location:", destination_input)
        print("Input OEM:", consignor)
        print("Matched Destination:", final_destination)
        print("Matched Consignee:", final_consignee)
        print("Matched OEM:", api_oem)
        print("OEM Score:", f"{oem_score:.2f}")
        print("Place Score:", f"{p_score:.2f}")
        print("Name Score:", f"{n_score:.2f}")
        print("Combined:", f"{combined:.2f}")
        print("Match Type:", match_type)
        print("Used Fallback:", used_fallback)
        print("Description:", description)
        print("MM Material:", mm_material)
        print("EV:", ev_flag)
        print("Rate:", rate)
        print("Freight:", freight)
        ws.cell(excel_row, 1).value = cn_no
        ws.cell(excel_row, 2).value = FIXED_CN_TYPE
        ws.cell(excel_row, 3).value = cn_date
        ws.cell(excel_row, 4).value = FIXED_OFFICE
        ws.cell(excel_row, 5).value = FIXED_BILLING_OFFICE
        ws.cell(excel_row, 6).value = consignor
        ws.cell(excel_row, 7).value = route
        ws.cell(excel_row, 8).value = route
        ws.cell(excel_row, 9).value = final_consignee
        ws.cell(excel_row, 10).value = trailer_no
        ws.cell(excel_row, 11).value = rate_chart
        ws.cell(excel_row, 12).value = load_type
        ws.cell(excel_row, 13).value = ""
        ws.cell(excel_row, 14).value = invoice_no
        ws.cell(excel_row, 15).value = invoice_date
        ws.cell(excel_row, 16).value = mm_material
        ws.cell(excel_row, 17).value = FIXED_WEIGHT
        ws.cell(excel_row, 18).value = rate
        ws.cell(excel_row, 19).value = freight
        ws.cell(excel_row, 20).value = ""
        ws.cell(excel_row, 21).value = chassis_no
        ws.cell(excel_row, 22).value = shipment_no
        ws.cell(excel_row, 23).value = ""
        ws.cell(excel_row, 24).value = ""
        ws.cell(excel_row, 25).value = ""
        ws.cell(excel_row, 26).value = ""
        if used_fallback:
            ws.cell(excel_row, 9).fill = RED_FILL
        excel_row += 1
        written_count += 1
    run_datetime = datetime.now().strftime("%Y-%m-%d_%H%M%S")
    output_file = f"TATA_SANAND_CN_{run_datetime}.xlsx"
    wb.save(output_file)
    archive_old_daily_files()
    final_output_path = os.path.join(DAILY_REPORTS_DIR, output_file)
    shutil.move(output_file, final_output_path)
    print("=" * 60)
    print(f"Saved to DailyReports: {final_output_path}")
    print(f"Rows written: {written_count}")
    print(f"Rows skipped because CN already exists in API: {skipped_count}")
    print("=" * 60)
    return final_output_path

if __name__ == "__main__":
    input_file = sys.argv[1] if len(sys.argv) > 1 else INPUT_FILE_DEFAULT
    filter_date = sys.argv[2] if len(sys.argv) > 2 else None
    result = process_tata_cn(input_file, filter_date=filter_date)
    print("Output file generated:", result)