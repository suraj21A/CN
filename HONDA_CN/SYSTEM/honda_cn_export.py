import os
import re
import sys
import time
import shutil
from difflib import SequenceMatcher
from datetime import datetime
import pandas as pd
import requests
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

TOOL_DIR = os.path.dirname(os.path.abspath(__file__))
ROOT_DIR = os.path.dirname(TOOL_DIR)

TEMPLATE_FILE = os.path.join(TOOL_DIR, "CN Bulk Upload Template.xlsx")
MASTER_FILE = os.path.join(TOOL_DIR, "HONDA_01.10.24.xlsx")

DAILY_REPORTS_DIR = os.path.join(ROOT_DIR, "DailyReports")
REPORT_REPOSITORY_DIR = os.path.join(ROOT_DIR, "Report Repository")

FIXED_CN_TYPE = "Billed"
FIXED_OFFICE = "Tapukara"
FIXED_BILLING_OFFICE = "Jaipur"
FIXED_CONSIGNOR = "Honda Cars India Limited"
FIXED_RATE_CHART = "Honda Rate Chart"
FIXED_WEIGHT = 1

EXPORT_PREFIX = "HE"
EXPORT_DESTINATION = "Pipavav"
EXPORT_CONSIGNEE = "Honda Cars India Limited"
EXPORT_FIXED_RATE = 17340.05

NAME_THRESHOLD = 0.80
PLACE_THRESHOLD = 0.85
OEM_THRESHOLD = 0.70

RED_FILL = PatternFill(fill_type="solid", start_color="FF9999", end_color="FF9999")

PLACE_ALIASES = {
    "BANGALORE": "BENGALURU",
    "BENGALORE": "BENGALURU",
    "GURGAON": "GURUGRAM",
    "BOMBAY": "MUMBAI",
    "CALCUTTA": "KOLKATA",
    "TRIVANDRUM": "THIRUVANANTHAPURAM",
    "ANANTHAPUR": "ANANTAPUR",
    "HARDWAR": "HARIDWAR",
    "HISSAR": "HISAR",
    "AMARAVATI": "AMRAVATI",
    "VADODARA": "BARODA",
    "BARODA": "BARODA",
    "BADODA": "BARODA",
    "BRODA": "BARODA",
    "SHOLAPUR": "SOLAPUR",
    "NEWDELHI": "DELHI",
    "NEW DELHI": "DELHI",
    "PIPAVAV PORT": "PIPAVAV",
    "PIPAVAV": "PIPAVAV",
}

CHARGED_ROUTE_OVERRIDES = {
    "ALAPPUZHA": "THRISSUR"
}
# --- OVERRIDE mapping, fill this as per your needs ---
OVERRIDES = {
    # ("DEALER NAME", "CITY"): ("CORRECTED DEALER", "CORRECTED CITY"),
}


def clean_text(val):
    if pd.isna(val):
        return ""
    return str(val).strip()

def clean_key(val):
    if pd.isna(val):
        return ""
    return str(val).strip().upper().replace(" ", "")

def parse_date(value):
    if pd.isna(value):
        return ""
    try:
        dt = pd.to_datetime(value, format="%Y-%m-%d", errors="coerce")
        if pd.isna(dt):
            dt = pd.to_datetime(value, dayfirst=True, errors="coerce")
        if pd.isna(dt):
            return str(value).strip()
        return dt.strftime("%Y-%m-%d")
    except Exception:
        return str(value).strip()

def normalize_name(name):
    if pd.isna(name):
        return ""
    text = str(name).upper().strip()
    text = re.sub(r"\bPRIVATE\s+LIMITED\b", " PVTLTD ", text)
    text = re.sub(r"\bPVT\.?\s*LTD\.?\b", " PVTLTD ", text)
    text = re.sub(r"\bPVT\s+LIMITED\b", " PVTLTD ", text)
    text = re.sub(r"\bPRIVATE\b", " PVT ", text)
    text = re.sub(r"\bLIMITED\b", " LTD ", text)
    text = re.sub(r"\bLLP\b", " LLP ", text)
    text = re.sub(r"[^A-Z0-9]", "", text)
    return text

def normalize_place(place):
    if pd.isna(place):
        return ""
    raw = str(place).upper().strip()
    raw = raw.replace(".", "").replace(",", "")
    raw = " ".join(raw.split())
    compact = re.sub(r"[^A-Z0-9]", "", raw)
    if raw in PLACE_ALIASES:
        alias = PLACE_ALIASES[raw]
        return re.sub(r"[^A-Z0-9]", "", alias.upper())
    if compact in PLACE_ALIASES:
        alias = PLACE_ALIASES[compact]
        return re.sub(r"[^A-Z0-9]", "", alias.upper())
    return compact

def similarity(a, b):
    if not a or not b:
        return 0.0
    if a == b:
        return 1.0
    if a in b or b in a:
        return 0.96
    return SequenceMatcher(None, a, b).ratio()

def name_similarity(a, b):
    return similarity(normalize_name(a), normalize_name(b))

def place_similarity(a, b):
    return similarity(normalize_place(a), normalize_place(b))

def get_root_name(name):
    if not name:
        return ""
    text = str(name).upper().strip().replace(".", " ")
    remove_words = {
        "MOTORS", "AUTOMOBILES", "AUTOMOBILE", "HONDA",
        "CARS", "CAR", "PVT", "PRIVATE", "LIMITED",
        "LLP", "LTD", "PVTLTD", "MOTOR", "AUTO"
    }
    parts = [p for p in text.split() if p not in remove_words]
    return parts[0] if parts else ""

def find_override_match(input_consignee, input_city, input_oem):
    norm_consignee = normalize_name(input_consignee)
    norm_city = normalize_place(input_city)
    for (k_name, k_city), (o_name, o_city) in OVERRIDES.items():
        if (
            norm_consignee == normalize_name(k_name)
            and norm_city == normalize_place(k_city)
        ):
            return {
                "geoName": o_name,
                "placeName": o_city,
                "dealerOEM": input_oem,
                "oem_score": 1.0,
                "used_original_fallback": False,
                "match_type": "OVERRIDE",
            }
    return None

def find_root_destination_match(input_consignee, input_city, input_oem, papl_data):
    root_input = get_root_name(input_consignee)
    city_norm_input = normalize_place(input_city)
    if not root_input or not city_norm_input:
        return None
    for area in papl_data:
        geo = str(area.get("geoName", ""))
        place = str(area.get("placeName", ""))
        dealer_oem = str(area.get("dealerOEM", ""))
        if not geo or not place:
            continue
        root_api = get_root_name(geo)
        city_norm_api = normalize_place(place)
        oem_score = name_similarity(input_oem, dealer_oem)
        if (root_api and root_input == root_api and city_norm_input == city_norm_api and oem_score >= OEM_THRESHOLD):
            return {
                "geoName": geo,
                "placeName": place,
                "dealerOEM": dealer_oem,
                "place_score": 1.0,
                "name_score": 1.0,
                "oem_score": oem_score,
                "combined": 1.0,
                "used_original_fallback": False,
                "match_type": "ROOT_NAME_PLACE_OEM",
            }
    return None

def find_best_papl_match(input_consignee, input_city, input_oem, papl_data):
    override = find_override_match(input_consignee, input_city, input_oem)
    if override:
        return override
    root_match = find_root_destination_match(input_consignee, input_city, input_oem, papl_data)
    if root_match:
        return root_match
    candidates = []
    for area in papl_data:
        geo = str(area.get("geoName", ""))
        place = str(area.get("placeName", ""))
        dealer_oem = str(area.get("dealerOEM", ""))
        if not geo or not place:
            continue
        p_score = place_similarity(input_city, place)
        n_score = name_similarity(input_consignee, geo)
        o_score = name_similarity(input_oem, dealer_oem)
        combined = (p_score * 0.30) + (n_score * 0.40) + (o_score * 0.30)
        candidates.append({
            "geoName": geo,
            "placeName": place,
            "dealerOEM": dealer_oem,
            "place_score": p_score,
            "name_score": n_score,
            "oem_score": o_score,
            "combined": combined,
            "used_original_fallback": False,
            "match_type": "FUZZY",
        })
    place_matched = [
        c for c in candidates 
        if c["place_score"] >= PLACE_THRESHOLD and c["oem_score"] >= OEM_THRESHOLD
    ]
    if place_matched:
        best = max(place_matched, key=lambda x: (x["name_score"], x["combined"]))
        if best["name_score"] >= NAME_THRESHOLD:
            return best
    if candidates:
        best = max(candidates, key=lambda x: x["combined"])
        if (best["name_score"] >= NAME_THRESHOLD and 
            best["place_score"] >= 0.75 and
            best["oem_score"] >= 0.70):
            return best
    return {
        "geoName": input_consignee,
        "placeName": input_city,
        "dealerOEM": input_oem,
        "oem_score": 0.0,
        "used_original_fallback": True,
        "match_type": "ORIGINAL_INPUT",
    }

def get_papl_data():
    url = "http://82.25.108.99:8375/retrieveAllPolygonAreas"
    try:
        response = requests.get(url, timeout=10)
        if response.status_code == 200:
            data = response.json()
            return data if isinstance(data, list) else []
        return []
    except Exception as e:
        print(f"Polygon API failed: {e}")
        return []

def extract_car_count(carrier_type):
    text = clean_text(carrier_type).upper()
    match = re.search(r"(\d+)", text)
    return match.group(1) if match else ""

def build_vehicle_load_count_map(df):
    vehicle_series = df["First Mile Carrier"].apply(lambda x: clean_text(x).upper())
    vehicle_series = vehicle_series[vehicle_series != ""]
    return vehicle_series.value_counts().to_dict()

def build_load_type_by_vehicle(vehicle_no, vehicle_count_map):
    vehicle_key = clean_text(vehicle_no).upper()
    count = vehicle_count_map.get(vehicle_key, 0)
    if count:
        return f"Honda Fixed {count} Car"
    return "Honda Fixed"

def normalize_honda_model(model_value):
    text = clean_text(model_value).upper()
    if "AMAZE" in text:
        return "AMAZE"
    if "ELEVATE" in text:
        return "ELEVATE"
    if "CITY" in text:
        return "CITY"
    if "WR-V" in text or "WRV" in text:
        return "WRV"
    if "JAZZ" in text:
        return "JAZZ"
    if "CIVIC" in text:
        return "CIVIC"
    if "BRIO" in text:
        return "BRIO"
    if "ACCORD" in text:
        return "ACCORD"
    if "CR-V" in text or "CRV" in text:
        return "CRV"
    return clean_text(model_value)

def build_cn_numbers(df):
    cn_numbers = []
    prev_vehicle = None
    vehicle_cn_cities = {}
    vehicle_cn_city_suffix = {}

    export_prefix = "HE"  # Update if your export prefix is different

    for idx, row in df.iterrows():
        vehicle = clean_text(row.get("First Mile Carrier")).upper()
        base_cn = clean_text(row.get("Invoice Number"))
        city = clean_text(row.get("City")).upper()
        dealer_code = clean_text(row.get("Dealer Code")).upper()

        # EXPORT/PIPAVAV FIX -- always no suffix
        is_export = dealer_code.startswith(export_prefix) or city == "PIPAVAV"
        if is_export:
            cn_numbers.append(base_cn)
            prev_vehicle = vehicle
            # Reset counters for next vehicle, if any
            vehicle_cn_cities = {}
            vehicle_cn_city_suffix = {}
            continue

        # New vehicle? Reset mappings
        if vehicle != prev_vehicle:
            vehicle_cn_cities = {}
            vehicle_cn_city_suffix = {}
            prev_vehicle = vehicle

        # How many unique CN numbers in this vehicle?
        this_vehicle_rows = df[df["First Mile Carrier"].apply(lambda x: clean_text(x).upper() == vehicle)]
        all_cns = this_vehicle_rows["Invoice Number"].apply(clean_text).unique()

        if len(all_cns) > 1:
            # CASE: Multiple different CN numbers in this vehicle - all as-is, no suffix, even if same
            cn_numbers.append(base_cn)
            continue

        # For this CN, track unique cities in this vehicle
        if base_cn not in vehicle_cn_cities:
            vehicle_cn_cities[base_cn] = set()
        vehicle_cn_cities[base_cn].add(city)

        if len(vehicle_cn_cities[base_cn]) == 1:
            # Only one city for this CN number in this vehicle, no suffix
            cn_numbers.append(base_cn)
            continue
        else:
            # Multiple cities for this CN, assign city-wise suffix
            if base_cn not in vehicle_cn_city_suffix:
                vehicle_cn_city_suffix[base_cn] = {}
            if city not in vehicle_cn_city_suffix[base_cn]:
                suffix = f"-{len(vehicle_cn_city_suffix[base_cn])}"
                vehicle_cn_city_suffix[base_cn][city] = suffix
            cn_numbers.append(f"{base_cn}{vehicle_cn_city_suffix[base_cn][city]}")
    return cn_numbers

def archive_old_daily_files():
    os.makedirs(DAILY_REPORTS_DIR, exist_ok=True)
    os.makedirs(REPORT_REPOSITORY_DIR, exist_ok=True)
    today_folder = datetime.now().strftime("%Y-%m-%d")
    repo_subfolder = os.path.join(REPORT_REPOSITORY_DIR, today_folder)
    os.makedirs(repo_subfolder, exist_ok=True)
    existing_files = [
        f for f in os.listdir(DAILY_REPORTS_DIR)
        if f.lower().endswith(".xlsx")
    ]
    for old_file in existing_files:
        old_file_path = os.path.join(DAILY_REPORTS_DIR, old_file)
        file_name_wo_ext, ext = os.path.splitext(old_file)
        timestamp = datetime.now().strftime("%H%M%S")
        dest_path = os.path.join(repo_subfolder, old_file)
        if os.path.exists(dest_path):
            dest_path = os.path.join(
                repo_subfolder,
                f"{file_name_wo_ext}_{timestamp}{ext}"
            )
        shutil.move(old_file_path, dest_path)

def prepare_input_df(df):
    df.columns = [str(c).strip() for c in df.columns]
    required_cols = [
        "Dealer Code", "Dealer Name", "City", "Model", "AR Invoice", "Invoice Date", "Invoice Number",
        "First Mile Carrier", "Carrier Type", "Dispatch Date", "Chasis Number"
    ]
    missing = [c for c in required_cols if c not in df.columns]
    if missing:
        raise Exception(f"Missing required columns in Honda input file: {missing}")
    df = df.dropna(how="all").copy()
    df = df[df["Invoice Number"].notna()].copy()
    df = df.reset_index(drop=True)
    return df

def detect_master_sheet(xls):
    preferred = []
    for s in xls.sheet_names:
        s_clean = s.strip().lower()
        if "table" in s_clean or "rate" in s_clean or "sheet" in s_clean:
            preferred.append(s)
    if preferred:
        return preferred[0]
    return xls.sheet_names[0]

def prepare_master_df(master_path):
    xls = pd.ExcelFile(master_path)
    sheet_name = detect_master_sheet(xls)
    raw = pd.read_excel(master_path, sheet_name=sheet_name, header=None)
    header_row_idx = None
    for idx in range(min(20, len(raw))):
        row_vals = [str(v).strip().upper() for v in raw.iloc[idx].tolist()]
        if "FROM" in row_vals and "TO" in row_vals:
            header_row_idx = idx
            break
    if header_row_idx is None:
        raise Exception("Could not identify master header row with From / To / TK4 / TR6 / TR7 / TR8")
    master = pd.read_excel(master_path, sheet_name=sheet_name, header=header_row_idx)
    master.columns = [str(c).strip() for c in master.columns]
    master = master.dropna(how="all").copy()
    required_cols = ["From", "To", "TK4", "TR6", "TR7", "TR8"]
    missing = [c for c in required_cols if c not in master.columns]
    if missing:
        raise Exception(f"Missing required columns in master file: {missing}")
    master["to_key"] = master["To"].apply(clean_key)
    return master

def get_rate_from_master(destination, carrier_type, master_df):
    destination_key = clean_key(destination)
    carrier_col = clean_text(carrier_type).upper()
    if not destination_key or not carrier_col:
        return None
    if carrier_col not in ["TK4", "TR6", "TR7", "TR8"]:
        return None

    # --- 1: Try exact match first ---
    match = master_df[master_df["to_key"] == destination_key]
    if not match.empty:
        rate = pd.to_numeric(match.iloc[0][carrier_col], errors="coerce")
        if pd.notna(rate):
            return float(rate)

    # --- 2: Fuzzy destination match ---
    max_sim = 0
    fuzzy_rate = None
    for idx, row in master_df.iterrows():
        sim = place_similarity(destination, row.get("To", ""))
        if sim > max_sim and sim >= 0.75:
            try:
                candidate_rate = float(row.get(carrier_col, 1))
                if not pd.isna(candidate_rate):
                    max_sim = sim
                    fuzzy_rate = candidate_rate
            except:
                continue
    if max_sim >= 0.75 and fuzzy_rate is not None:
        print(f"Fuzzy rate match for dest: {destination}, sim: {max_sim:.2f}, rate: {fuzzy_rate}")
        return fuzzy_rate

    print(f"Destination not found in master (even fuzzy): {destination}")
    return None
def process_honda_cn(input_file, filter_date=None):
    print("=" * 60)
    print("HONDA CN PROCESS STARTED")
    print("=" * 60)
    print("Input file:", input_file)
    print("Master file:", MASTER_FILE)
    df = pd.read_excel(input_file)
    df = prepare_input_df(df)
    master_df = prepare_master_df(MASTER_FILE)
    papl_data = get_papl_data()
    if filter_date:
        df["parsed_dispatch_date"] = pd.to_datetime(df["Dispatch Date"], errors="coerce", dayfirst=True)
        filter_dt = pd.to_datetime(filter_date, errors="coerce")
        if pd.notna(filter_dt):
            df = df[df["parsed_dispatch_date"].dt.date == filter_dt.date()]

    if df.empty:
        print("No data found after applying filter.")
        return None

    df = df.reset_index(drop=True)
    df["FINAL_CN_NO"] = build_cn_numbers(df)
    vehicle_count_map = build_vehicle_load_count_map(df)
    wb = load_workbook(TEMPLATE_FILE)
    ws = wb.active

    for row_idx in range(2, ws.max_row + 1):
        for col in range(1, 27):
            ws.cell(row_idx, col).value = None
            ws.cell(row_idx, col).fill = PatternFill(fill_type=None)

    for i, (_, row) in enumerate(df.iterrows(), start=2):
        cn_no = clean_text(row.get("FINAL_CN_NO"))
        dealer_code = clean_text(row.get("Dealer Code"))
        dealer_name = clean_text(row.get("Dealer Name"))
        city = clean_text(row.get("City"))
        model_raw = clean_text(row.get("Model"))
        model = normalize_honda_model(model_raw)
        ar_invoice = clean_text(row.get("AR Invoice"))
        invoice_date = parse_date(row.get("Invoice Date"))
        invoice_number = clean_text(row.get("Invoice Number"))
        vehicle_no = clean_text(row.get("First Mile Carrier"))
        carrier_type = clean_text(row.get("Carrier Type")).upper()
        dispatch_date = parse_date(row.get("Dispatch Date"))
        chassis_no = clean_text(row.get("Chasis Number"))
        is_export = dealer_code.upper().startswith(EXPORT_PREFIX)

        if is_export:
            destination = EXPORT_DESTINATION
            consignee = EXPORT_CONSIGNEE
            match_type = "EXPORT_FIXED"
            used_fallback = False
            oem_score = 1.0
            api_oem = EXPORT_CONSIGNEE
            rate = EXPORT_FIXED_RATE
            freight = EXPORT_FIXED_RATE
        else:
            papl_match = find_best_papl_match(dealer_name, city, FIXED_CONSIGNOR, papl_data)
            destination = papl_match["placeName"]
            consignee = papl_match["geoName"]
            match_type = papl_match.get("match_type")
            used_fallback = papl_match.get("used_original_fallback", False)
            api_oem = papl_match.get("dealerOEM", "")
            oem_score = papl_match.get("oem_score", 0.0)
            rate = get_rate_from_master(destination, carrier_type, master_df)
            if rate is None:
                rate = 1
            freight = rate

        mm_invoice_no = ar_invoice if ar_invoice else invoice_number
        load_type = build_load_type_by_vehicle(vehicle_no, vehicle_count_map)
        actual_route = f"{FIXED_OFFICE}-{destination}-{FIXED_CONSIGNOR}"
        charged_route = actual_route

        print("-" * 60)
        print(f"Row {i}")
        print("Final CN No:", cn_no)
        print("Dealer Code:", dealer_code)
        print("Dealer Name:", dealer_name)
        print("City:", city)
        print("Model Raw:", model_raw)
        print("MM Material Final:", model)
        print("Export:", is_export)
        print("Input OEM:", FIXED_CONSIGNOR)
        print("Final Destination:", destination)
        print("Final Consignee:", consignee)
        print("Matched OEM:", api_oem)
        print("OEM Score:", f"{oem_score:.2f}")
        print("Match Type:", match_type)
        print("Fallback Used:", used_fallback)
        print("Invoice Number:", invoice_number)
        print("MM Invoice No:", mm_invoice_no)
        print("Vehicle No:", vehicle_no)
        print("Carrier Type:", carrier_type)
        print("Load Type:", load_type)
        print("Rate Final:", rate)
        print("Freight Final:", freight)

        ws.cell(i, 1).value = cn_no
        ws.cell(i, 2).value = FIXED_CN_TYPE
        ws.cell(i, 3).value = dispatch_date
        ws.cell(i, 4).value = FIXED_OFFICE
        ws.cell(i, 5).value = FIXED_BILLING_OFFICE
        ws.cell(i, 6).value = FIXED_CONSIGNOR
        ws.cell(i, 7).value = actual_route
        ws.cell(i, 8).value = charged_route
        ws.cell(i, 9).value = consignee
        ws.cell(i, 10).value = vehicle_no
        ws.cell(i, 11).value = FIXED_RATE_CHART
        ws.cell(i, 12).value = load_type
        ws.cell(i, 13).value = ""
        ws.cell(i, 14).value = mm_invoice_no
        ws.cell(i, 15).value = invoice_date
        ws.cell(i, 16).value = model
        ws.cell(i, 17).value = FIXED_WEIGHT
        ws.cell(i, 18).value = rate
        ws.cell(i, 19).value = freight
        ws.cell(i, 20).value = ""
        ws.cell(i, 21).value = chassis_no
        ws.cell(i, 22).value = ""
        ws.cell(i, 23).value = dealer_code
        ws.cell(i, 24).value = carrier_type
        ws.cell(i, 25).value = ""
        ws.cell(i, 26).value = ""

        if (not is_export) and used_fallback:
            ws.cell(i, 9).fill = RED_FILL

    run_datetime = datetime.now().strftime("%Y-%m-%d_%H%M%S")
    output_file = f"HONDA_CN{run_datetime}.xlsx"

    wb.save(output_file)
    archive_old_daily_files()
    final_output_path = os.path.join(DAILY_REPORTS_DIR, output_file)
    shutil.move(output_file, final_output_path)
    print("=" * 60)
    print(f"Saved to DailyReports: {final_output_path}")
    print("=" * 60)
    return final_output_path

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python honda_cn_export.py <input_file.xlsx> [YYYY-MM-DD]")
        sys.exit(1)
    input_file = sys.argv[1]
    filter_date = sys.argv[2] if len(sys.argv) > 2 else None
    result = process_honda_cn(input_file, filter_date=filter_date)
    print("Output file generated:", result)