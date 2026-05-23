import os
import shutil
import time
import sys
import re
from difflib import SequenceMatcher
from datetime import datetime
import pandas as pd
import requests
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

TOOL_DIR = os.path.dirname(os.path.abspath(__file__))
ROOT_DIR = os.path.dirname(TOOL_DIR)

TEMPLATE_FILE = os.path.join(TOOL_DIR, "CN Bulk Upload Template.xlsx")
DAILY_REPORTS_DIR = os.path.join(ROOT_DIR, "DailyReports")
REPORT_REPOSITORY_DIR = os.path.join(ROOT_DIR, "Report Repository")

FIXED_CONSIGNOR = "GLOVIS INDIA PVT. LTD."
FIXED_CN_TYPE = "Billed"
FIXED_OFFICE = "Pune"
FIXED_BILLING_OFFICE = "Pune"
FIXED_RATE_CHART = "Hyundai Rate Chart"
FIXED_LOAD_TYPE = "Hyundai Fixed"
FIXED_WEIGHT = 1

NAME_THRESHOLD = 0.80
PLACE_THRESHOLD = 0.85
OEM_THRESHOLD = 0.70  # Minimum OEM match score threshold

RED_FILL = PatternFill(fill_type="solid", start_color="FF9999", end_color="FF9999")

MODEL_MAP = {
    "HQ S4": "Aura",
    "HQ W5": "Exter",
    "FH W5": "Creta",
    "7H W5": "Creta",
    "FH WB": "Creta",
    "SV S6": "i20",
    "0Y S4": "Verna",
    "FH WC": "Alcazar",
    "HQ S6": "Grand",
    "HO W5": "Kona",
    "9M WD": "Tucson",
    "6I WS": "Ioniq 5",
    "1Q W5": "Venue",

}

PLACE_ALIASES = {
    "BANGALORE": "BENGALURU",
    "BENGALORE": "BENGALURU",
    "GURGAON": "GURUGRAM",
    "BOMBAY": "MUMBAI",
    "CALCUTTA": "KOLKATA",
    "TRIVANDRUM": "THIRUVANANTHAPURAM",
    "ANANTHAPUR": "ANANTAPUR",
    "HARDWAR": "HARIDWAR",
    "HISSAR": "HISAR",
    "AMARAVATI": "AMRAVATI",
    "VADODARA": "BARODA",
    "BARODA": "BARODA",
    "BADODA": "BARODA",
    "SHOLAPUR": "SOLAPUR",
    "NEWDELHI": "DELHI",
    "NEW DELHI": "DELHI",
    "REDHILLSCHENNAI": "CHENNAI",
    "REDHILLS-CHENNAI": "CHENNAI",
}


# ⚡️ Add OVERRIDE MAPPING for Hyundai
OVERRIDES = {
    ("ABC Hyundai Dealers Pvt Ltd", "Delhi"): ("ABC Hyundai Dealers Pvt Ltd", "DELHI"),
    # Add all necessary (dealer, city): (override_dealer, override_city) pairs here
    # Example: ("ORIGINAL NAME", "CITY"): ("CORRECTED NAME", "CORRECTED CITY")
}

def clean_text(val):
    if pd.isna(val):
        return ""
    return str(val).strip()

def clean_key(val):
    if pd.isna(val):
        return ""
    return str(val).strip().upper().replace(" ", "")

def parse_date(value):
    if pd.isna(value):
        return ""
    try:
        dt = pd.to_datetime(value, errors="coerce", dayfirst=True)
        if pd.isna(dt):
            return str(value).strip()
        return dt.strftime("%Y-%m-%d")
    except Exception:
        return str(value).strip()

def normalize_model_code(model_value):
    if pd.isna(model_value):
        return ""
    model_text = str(model_value).strip().upper()
    return MODEL_MAP.get(model_text, str(model_value).strip())

def make_base_cn_no(lot_no):
    """
    Example:
    TP041FQ91920260321503 -> HYD20260321503
    """
    s = clean_text(lot_no)
    idx = s.find("2026")
    if idx == -1:
        return f"HYD{s}"
    return f"HYD{s[idx:]}"

# ---------------- API ----------------
def get_papl_data():
    url = "http://82.25.108.99:8375/retrieveAllPolygonAreas"
    try:
        response = requests.get(url, timeout=10)
        if response.status_code == 200:
            data = response.json()
            return data if isinstance(data, list) else []
        return []
    except Exception as e:
        print(f"Polygon API failed: {e}")
        return []

# ---------------- NORMALIZATION ----------------
def normalize_name(name):
    if pd.isna(name):
        return ""

    text = str(name).upper().strip()
    text = re.sub(r"\bPRIVATE\s+LIMITED\b", " PVTLTD ", text)
    text = re.sub(r"\bPVT\.?\s*LTD\.?\b", " PVTLTD ", text)
    text = re.sub(r"\bPVT\s+LIMITED\b", " PVTLTD ", text)
    text = re.sub(r"\bPRIVATE\b", " PVT ", text)
    text = re.sub(r"\bLIMITED\b", " LTD ", text)
    text = re.sub(r"\bLLP\b", " LLP ", text)
    text = re.sub(r"[^A-Z0-9]", "", text)
    return text

def normalize_place(place):
    if pd.isna(place):
        return ""
    raw = str(place).upper().strip()
    raw = raw.replace(".", "").replace(",", "")
    raw = " ".join(raw.split())
    compact = re.sub(r"[^A-Z0-9]", "", raw)
    if raw in PLACE_ALIASES:
        alias = PLACE_ALIASES[raw]
        return re.sub(r"[^A-Z0-9]", "", alias.upper())
    if compact in PLACE_ALIASES:
        alias = PLACE_ALIASES[compact]
        return re.sub(r"[^A-Z0-9]", "", alias.upper())
    return compact

def similarity(a, b):
    if not a or not b:
        return 0.0
    if a == b:
        return 1.0
    if a in b or b in a:
        return 0.96
    return SequenceMatcher(None, a, b).ratio()

def name_similarity(a, b):
    return similarity(normalize_name(a), normalize_name(b))

def place_similarity(a, b):
    return similarity(normalize_place(a), normalize_place(b))

# ⚡️ Root name extraction for fuzzy/strong logic
def get_root_name(name):
    """Extract root part of company name (removes generic suffix words)"""
    if not name:
        return ""
    text = str(name).upper().strip().replace(".", " ")
    remove_words = {
        "MOTORS", "AUTOMOBILES", "AUTOMOBILE", "CARS", "CAR", "PVT", "PRIVATE", "LIMITED", "LTD", "LLP", "PVTLTD", "AUTO", "AUTOMOTIVE",
        "WHEELS", "MOTOCORP", "KIOSK", "HYUNDAI", "HUNDAI"
    }
    parts = [p for p in text.split() if p not in remove_words and p]
    return parts[0] if parts else ""

# ⚡️ OVERRIDE checking logic
def find_override_match(input_consignee, input_city, input_oem):
    """Check for an exact override mapping"""
    norm_consignee = normalize_name(input_consignee)
    norm_city = normalize_place(input_city)
    for (k_name, k_city), (o_name, o_city) in OVERRIDES.items():
        if (
            norm_consignee == normalize_name(k_name)
            and norm_city == normalize_place(k_city)
        ):
            return {
                "geoName": o_name,
                "placeName": o_city,
                "dealerOEM": input_oem,
                "oem_score": 1.0,
                "used_original_fallback": False,
                "match_type": "OVERRIDE",
            }
    return None

# ⚡️ Full consginee fuzzy logic (Kia style)
def find_best_papl_match(input_consignee, input_city, input_oem, papl_data):
    # Step 1: OVERRIDE TABLE
    override = find_override_match(input_consignee, input_city, input_oem)
    if override:
        return override

    # Step 2: ROOT NAME + PLACE + OEM exact check
    root_input = get_root_name(input_consignee)
    city_norm_input = normalize_place(input_city)

    if root_input:
        for area in papl_data:
            geo = area.get("geoName")
            place = area.get("placeName")
            dealer_oem = area.get("dealerOEM", "")

            if not geo or not place:
                continue

            root_api = get_root_name(geo)
            city_norm_api = normalize_place(place)
            oem_match = normalize_name(input_oem) == normalize_name(dealer_oem)
            oem_score = 1.0 if oem_match else 0.0

            if (root_api and 
                root_input == root_api and 
                city_norm_input == city_norm_api and 
                oem_score >= OEM_THRESHOLD):

                return {
                    "geoName": geo,
                    "placeName": place,
                    "dealerOEM": dealer_oem,
                    "place_score": 1.0,
                    "name_score": 1.0,
                    "oem_score": oem_score,
                    "combined": 1.0,
                    "used_original_fallback": False,
                    "match_type": "ROOT_NAME_PLACE_OEM",
                }

    # Step 3: Normal fuzzy
    candidates = []
    for area in papl_data:
        geo = area.get("geoName")
        place = area.get("placeName")
        dealer_oem = area.get("dealerOEM", "")
        if not geo or not place:
            continue
        p_score = place_similarity(input_city, place)
        n_score = name_similarity(input_consignee, geo)
        o_score = name_similarity(input_oem, dealer_oem)
        combined = (p_score * 0.30) + (n_score * 0.40) + (o_score * 0.30)
        candidates.append({
            "geoName": geo,
            "placeName": place,
            "dealerOEM": dealer_oem,
            "place_score": p_score,
            "name_score": n_score,
            "oem_score": o_score,
            "combined": combined,
            "used_original_fallback": False,
            "match_type": "FUZZY",
        })

    # Fuzzy selection (thresholds)
    place_matched = [
        c for c in candidates 
        if c["place_score"] >= PLACE_THRESHOLD and c["oem_score"] >= OEM_THRESHOLD
    ]
    if place_matched:
        best = max(place_matched, key=lambda x: (x["name_score"], x["combined"]))
        if best["name_score"] >= NAME_THRESHOLD:
            return best

    if candidates:
        best = max(candidates, key=lambda x: x["combined"])
        if (best["name_score"] >= NAME_THRESHOLD and 
            best["place_score"] >= 0.75 and
            best["oem_score"] >= 0.70):
            return best

    return {
        "geoName": input_consignee,
        "placeName": input_city,
        "dealerOEM": input_oem,
        "place_score": 0.0,
        "name_score": 0.0,
        "oem_score": 0.0,
        "combined": 0.0,
        "used_original_fallback": True,
        "match_type": "ORIGINAL_INPUT",
    }


# ============ ARCHIVE SECTION ============
def archive_old_daily_files():
    os.makedirs(DAILY_REPORTS_DIR, exist_ok=True)
    os.makedirs(REPORT_REPOSITORY_DIR, exist_ok=True)

    today_folder = datetime.now().strftime("%Y-%m-%d")
    repo_subfolder = os.path.join(REPORT_REPOSITORY_DIR, today_folder)
    os.makedirs(repo_subfolder, exist_ok=True)

    existing_files = [
        f for f in os.listdir(DAILY_REPORTS_DIR)
        if f.lower().endswith(".xlsx")
    ]

    for old_file in existing_files:
        old_file_path = os.path.join(DAILY_REPORTS_DIR, old_file)
        file_name_wo_ext, ext = os.path.splitext(old_file)

        timestamp = datetime.now().strftime("%H%M%S")
        dest_path = os.path.join(repo_subfolder, old_file)

        if os.path.exists(dest_path):
            dest_path = os.path.join(
                repo_subfolder,
                f"{file_name_wo_ext}_{timestamp}{ext}"
            )

        shutil.move(old_file_path, dest_path)


# ============ DATA PREPARATION SECTION ============
def prepare_trpspn_df(trpspn_df):
    """Validate and clean input data from TRPSPN file"""
    trpspn_df.columns = [str(c).strip() for c in trpspn_df.columns]

    required_cols = [
        "Dealer Code",
        "Plate no",
        "Lot no",
        "VIN No",
        "Model",
        "Billing date",
        "Purch Amt",
    ]
    missing = [c for c in required_cols if c not in trpspn_df.columns]
    if missing:
        raise Exception(f"Missing required columns in trpspn file: {missing}")

    # Remove completely empty rows
    trpspn_df = trpspn_df.dropna(how="all").copy()

    # Filter rows with required data
    trpspn_df = trpspn_df[
        trpspn_df["Dealer Code"].notna() &
        trpspn_df["Plate no"].notna() &
        trpspn_df["Lot no"].notna()
    ].copy()

    # Create normalized keys for deduplication
    trpspn_df["dealer_key"] = trpspn_df["Dealer Code"].apply(clean_key)
    trpspn_df["plate_key"] = trpspn_df["Plate no"].apply(clean_key)
    trpspn_df["lot_key"] = trpspn_df["Lot no"].apply(clean_key)

    # Remove exact duplicate records
    trpspn_df = trpspn_df.drop_duplicates(
        subset=["dealer_key", "plate_key", "lot_key", "VIN No", "Model", "Billing date", "Purch Amt"],
        keep="first"
    ).reset_index(drop=True)

    return trpspn_df


def prepare_truck_df(truck_df):
    """Validate and clean support data from truck file"""
    truck_df.columns = [str(c).strip() for c in truck_df.columns]

    required_cols = [
        "DEALER_CD",
        "PLATE_NO",
        "GATE_OUT_DT",
        "HMI_LOT_NO",
        "LOCATION_NM",
        "DEALER_NM",
    ]
    missing = [c for c in required_cols if c not in truck_df.columns]
    if missing:
        raise Exception(f"Missing required columns in truck file: {missing}")

    # Remove completely empty rows
    truck_df = truck_df.dropna(how="all").copy()

    # Filter rows with required data
    truck_df = truck_df[
        truck_df["DEALER_CD"].notna() &
        truck_df["PLATE_NO"].notna() &
        truck_df["HMI_LOT_NO"].notna()
    ].copy()

    # Create normalized keys for merging
    truck_df["dealer_key"] = truck_df["DEALER_CD"].apply(clean_key)
    truck_df["plate_key"] = truck_df["PLATE_NO"].apply(clean_key)
    truck_df["lot_key"] = truck_df["HMI_LOT_NO"].apply(clean_key)

    # Keep only necessary columns
    truck_df = truck_df[
        ["dealer_key", "plate_key", "lot_key", "GATE_OUT_DT", "LOCATION_NM", "DEALER_NM"]
    ].copy()

    # Remove duplicates
    truck_df = truck_df.drop_duplicates(
        subset=["dealer_key", "plate_key", "lot_key"],
        keep="first"
    ).reset_index(drop=True)

    return truck_df


def merge_hyundai_data(trpspn_df, truck_df):
    """Merge TRPSPN and truck data using normalized keys"""
    merged_df = trpspn_df.merge(
        truck_df,
        on=["dealer_key", "plate_key", "lot_key"],
        how="left",
        validate="many_to_one"
    )
    return merged_df


# ============ CN NUMBER GENERATION SECTION ============
def build_cn_numbers(df):
    """
    Build CN numbers with dealer-based suffix logic.
    
    Rule:
    - New vehicle -> Use base CN (from lot number)
    - Same vehicle + same dealer -> Reuse same suffix
    - Same vehicle + new dealer -> Assign new suffix with increment
    
    Example:
    Vehicle 1 with Dealer A: HYD20260321503
    Vehicle 1 with Dealer B: HYD20260321503-0
    Vehicle 2 with Dealer A: HYD20260321504
    """
    cn_numbers = []
    prev_vehicle = None
    dealer_suffix_map = {}

    for _, row in df.iterrows():
        vehicle = clean_text(row["Plate no"]).upper()
        dealer = clean_text(row["Dealer Code"]).upper()
        base_cn = make_base_cn_no(row["Lot no"])

        if prev_vehicle is None or vehicle != prev_vehicle:
            # New vehicle - start fresh with no suffix
            cn_numbers.append(base_cn)
            prev_vehicle = vehicle
            dealer_suffix_map = {dealer: ""}
        else:
            # Same vehicle - check if dealer seen before
            if dealer in dealer_suffix_map:
                # Dealer seen for this vehicle - reuse suffix
                suffix = dealer_suffix_map[dealer]
                cn_numbers.append(f"{base_cn}{suffix}")
            else:
                # New dealer for this vehicle - create new suffix
                suffix = f"-{len(dealer_suffix_map)}"
                dealer_suffix_map[dealer] = suffix
                cn_numbers.append(f"{base_cn}{suffix}")

    return cn_numbers


# ============ MAIN PROCESSING SECTION ============
def process_hyundai_cn(trpspn_file, truck_file, filter_date=None):
    """
    Main processing function for Hyundai CN generation.
    Reads TRPSPN and truck files, merges data, matches to polygon areas via OEM, 
    and generates export file.
    """
    print("=" * 60)
    print("HYUNDAI CN PROCESS STARTED")
    print("=" * 60)
    print("TRPSPN file (base):", trpspn_file)
    print("Truck file (support):", truck_file)

    # Read and prepare input data
    trpspn_df = pd.read_excel(trpspn_file)
    truck_df = pd.read_excel(truck_file)

    trpspn_df = prepare_trpspn_df(trpspn_df)
    truck_df = prepare_truck_df(truck_df)

    print("TRPSPN rows after cleanup:", len(trpspn_df))
    print("Truck support rows after cleanup:", len(truck_df))

    # Merge support data
    merged_df = merge_hyundai_data(trpspn_df, truck_df)

    print("Merged rows:", len(merged_df))

    # Apply date filter if provided
    if filter_date:
        merged_df["parsed_gate_out"] = pd.to_datetime(
            merged_df["GATE_OUT_DT"], errors="coerce", dayfirst=True
        )
        filter_dt = pd.to_datetime(filter_date, errors="coerce")
        if pd.notna(filter_dt):
            merged_df = merged_df[merged_df["parsed_gate_out"].dt.date == filter_dt.date()]

    if merged_df.empty:
        print("No data found after applying filter.")
        return None

    # Build CN numbers
    merged_df = merged_df.reset_index(drop=True)
    merged_df["FINAL_CN_NO"] = build_cn_numbers(merged_df)

    # Fetch polygon area data from API
    papl_data = get_papl_data()

    # Load and clear template
    wb = load_workbook(TEMPLATE_FILE)
    ws = wb.active

    for row_idx in range(2, ws.max_row + 1):
        for col in range(1, 27):
            ws.cell(row_idx, col).value = None
            ws.cell(row_idx, col).fill = PatternFill(fill_type=None)

    # Process each row
    for i, (_, row) in enumerate(merged_df.iterrows(), start=2):
        dealer_code = clean_text(row["Dealer Code"])
        plate_no = clean_text(row["Plate no"])
        lot_no = clean_text(row["Lot no"])
        cn_no = clean_text(row["FINAL_CN_NO"])

        raw_consignee = clean_text(row.get("DEALER_NM"))
        raw_destination = clean_text(row.get("LOCATION_NM"))
        cn_date = parse_date(row.get("GATE_OUT_DT"))

        # Match via polygon API with OEM matching (3-factor scoring)
        papl_match = find_best_papl_match(raw_consignee, raw_destination, FIXED_CONSIGNOR, papl_data)
        final_consignee = papl_match["geoName"]
        final_destination = papl_match["placeName"]
        api_oem = papl_match.get("dealerOEM", "")  # Get matched OEM from API
        oem_score = papl_match.get("oem_score", 0.0)  # Get OEM match score

        vin_no = clean_text(row.get("VIN No"))
        model_code = clean_text(row.get("Model"))
        mm_material = normalize_model_code(model_code)
        billing_date = parse_date(row.get("Billing date"))

        # Get rate from purchase amount
        purch_amt = row.get("Purch Amt")
        rate = pd.to_numeric(purch_amt, errors="coerce")
        rate_value = "" if pd.isna(rate) else float(rate)
        freight_value = "" if pd.isna(rate) else float(rate)

        # Build route strings
        actual_route = f"{FIXED_OFFICE}-{final_destination}-{FIXED_CONSIGNOR}" if final_destination else f"{FIXED_OFFICE}--{FIXED_CONSIGNOR}"
        charged_route = actual_route

        # Log processing details
        print("-" * 60)
        print(f"Row {i}")
        print("Dealer Code:", dealer_code)
        print("Plate No:", plate_no)
        print("Lot No:", lot_no)
        print("CN No:", cn_no)
        print("Raw Consignee:", raw_consignee)
        print("Raw Destination:", raw_destination)
        print("Input OEM:", FIXED_CONSIGNOR)  # OEM used for matching
        print("Matched Consignee:", final_consignee)
        print("Matched Destination:", final_destination)
        print("Matched OEM:", api_oem)  # OEM from API match
        print("OEM Score:", f"{oem_score:.2f}")  # OEM matching score
        print("Match Type:", papl_match.get("match_type"))
        print("Fallback Used:", papl_match.get("used_original_fallback"))
        print("Billing Date:", billing_date)
        print("CN Date:", cn_date)
        print("VIN No:", vin_no)
        print("Model Code:", model_code)
        print("MM Material:", mm_material)
        print("Purchase Amt:", purch_amt)

        # Write to Excel
        ws.cell(i, 1).value = cn_no
        ws.cell(i, 2).value = FIXED_CN_TYPE
        ws.cell(i, 3).value = cn_date
        ws.cell(i, 4).value = FIXED_OFFICE
        ws.cell(i, 5).value = FIXED_BILLING_OFFICE
        ws.cell(i, 6).value = FIXED_CONSIGNOR
        ws.cell(i, 7).value = actual_route
        ws.cell(i, 8).value = charged_route
        ws.cell(i, 9).value = final_consignee
        ws.cell(i, 10).value = plate_no
        ws.cell(i, 11).value = FIXED_RATE_CHART
        ws.cell(i, 12).value = FIXED_LOAD_TYPE
        ws.cell(i, 13).value = ""
        ws.cell(i, 14).value = lot_no
        ws.cell(i, 15).value = billing_date
        ws.cell(i, 16).value = mm_material
        ws.cell(i, 17).value = FIXED_WEIGHT
        ws.cell(i, 18).value = rate_value
        ws.cell(i, 19).value = freight_value
        ws.cell(i, 20).value = ""
        ws.cell(i, 21).value = vin_no
        ws.cell(i, 22).value = ""
        ws.cell(i, 23).value = ""
        ws.cell(i, 24).value = ""
        ws.cell(i, 25).value = ""
        ws.cell(i, 26).value = ""

        # Highlight fallback matches in red
        if papl_match.get("used_original_fallback"):
            ws.cell(i, 9).fill = RED_FILL

    # Save and archive
    run_datetime = datetime.now().strftime("%Y-%m-%d_%H%M%S")
    output_file = f"HUNDAI_PUNE_CN_{run_datetime}.xlsx"

    wb.save(output_file)

    archive_old_daily_files()

    final_output_path = os.path.join(DAILY_REPORTS_DIR, output_file)
    shutil.move(output_file, final_output_path)
    print("=" * 60)
    print(f"Saved to DailyReports: {final_output_path}")
    print("=" * 60)

    return final_output_path


if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("Usage: python hyundai_cn_export.py <trpspn_file.xlsx> <truck_file.xlsx> [YYYY-MM-DD]")
        sys.exit(1)

    trpspn_file = sys.argv[1]
    truck_file = sys.argv[2]
    filter_date = sys.argv[3] if len(sys.argv) > 3 else None

    result = process_hyundai_cn(trpspn_file, truck_file, filter_date=filter_date)
    print("Output file generated:", result)