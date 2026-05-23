import os
import shutil
import time
import sys
import pandas as pd
from openpyxl import load_workbook
from datetime import datetime
from db_helper import save_cn_to_mysql
TOOL_DIR = os.path.dirname(os.path.abspath(__file__))
ROOT_DIR = os.path.dirname(TOOL_DIR)

TEMPLATE_FILE = os.path.join(TOOL_DIR, "CN Bulk Upload Template.xlsx")
DAILY_REPORTS_DIR = os.path.join(ROOT_DIR, "DailyReports")
REPORT_REPOSITORY_DIR = os.path.join(ROOT_DIR, "Report Repository")

FIXED_CONSIGNOR = "GLOVIS INDIA PVT. LTD."
FIXED_CONSIGNEE = "GLOVIS INDIA PVT. LTD."
FIXED_CN_TYPE = "Billed"
FIXED_OFFICE = "Chennai"
FIXED_BILLING_OFFICE = "Chennai"
FIXED_RATE_CHART = "Hyundai Rate Chart"
FIXED_LOAD_TYPE = "Hyundai Fixed"
FIXED_WEIGHT = 1
FIXED_DESTINATION = "Chennai Port"

MODEL_MAP = {
    "HQ:S4": "Aura",
    "HQ:W5": "Exter",
    "FH:W5": "Creta",
    "7H:W5": "Creta",
    "FH:WB": "Creta",
    "SV:S6": "i20",
    "0Y:S4": "Verna",
    "FH:WC": "Alcazar",
    "HQ:S6": "Grand",
    "HO:W5": "Kona",
    "9M:WD": "Tucson",
    "6I:WS": "Ioniq 5",
}


def clean_text(val):
    if pd.isna(val):
        return ""
    return str(val).strip()


def clean_key(val):
    if pd.isna(val):
        return ""
    return str(val).strip().upper().replace(" ", "")


def parse_date(value):
    if pd.isna(value):
        return ""
    try:
        dt = pd.to_datetime(value, errors="coerce", dayfirst=True)
        if pd.isna(dt):
            return str(value).strip()
        return dt.strftime("%Y-%m-%d")
    except Exception:
        return str(value).strip()


def normalize_model_code(model_value):
    if pd.isna(model_value):
        return ""
    model_text = str(model_value).strip().upper()
    return MODEL_MAP.get(model_text, str(model_value).strip())


def make_base_cn_no(lot_no):
    """
    Example:
    TP041FQ91920260321503 -> HYE20260321503
    """
    s = clean_text(lot_no)
    idx = s.find("2026")
    if idx == -1:
        return f"HYE{s}"
    return f"HYE{s[idx:]}"


def archive_old_daily_files():
    os.makedirs(DAILY_REPORTS_DIR, exist_ok=True)
    os.makedirs(REPORT_REPOSITORY_DIR, exist_ok=True)

    today_folder = datetime.now().strftime("%Y-%m-%d")
    repo_subfolder = os.path.join(REPORT_REPOSITORY_DIR, today_folder)
    os.makedirs(repo_subfolder, exist_ok=True)

    existing_files = [
        f for f in os.listdir(DAILY_REPORTS_DIR)
        if f.lower().endswith(".xlsx")
    ]

    for old_file in existing_files:
        old_file_path = os.path.join(DAILY_REPORTS_DIR, old_file)
        file_name_wo_ext, ext = os.path.splitext(old_file)

        timestamp = datetime.now().strftime("%H%M%S")
        dest_path = os.path.join(repo_subfolder, old_file)

        if os.path.exists(dest_path):
            dest_path = os.path.join(
                repo_subfolder,
                f"{file_name_wo_ext}_{timestamp}{ext}"
            )

        shutil.move(old_file_path, dest_path)


def prepare_trpspn_df(trpspn_df):
    trpspn_df.columns = [str(c).strip() for c in trpspn_df.columns]

    required_cols = [
        "Distributor Code",
        "Plate no",
        "Lot no",
        "VIN No",
        "Model",
        "Billing date",
        "Purch Amt",
        "Gate Out Date",
    ]
    missing = [c for c in required_cols if c not in trpspn_df.columns]
    if missing:
        raise Exception(f"Missing required columns in TRPSPN file: {missing}")

    trpspn_df = trpspn_df.dropna(how="all").copy()

    trpspn_df = trpspn_df[
        trpspn_df["Distributor Code"].notna() &
        trpspn_df["Plate no"].notna() &
        trpspn_df["Lot no"].notna()
    ].copy()

    trpspn_df["dealer_key"] = trpspn_df["Distributor Code"].apply(clean_key)
    trpspn_df["plate_key"] = trpspn_df["Plate no"].apply(clean_key)
    trpspn_df["lot_key"] = trpspn_df["Lot no"].apply(clean_key)

    trpspn_df = trpspn_df.drop_duplicates(
        subset=["dealer_key", "plate_key", "lot_key", "VIN No", "Model", "Billing date", "Purch Amt"],
        keep="first"
    ).reset_index(drop=True)

    return trpspn_df


def build_cn_numbers(df):
    """
    KIA-style suffix logic:
    - New vehicle -> base CN No
    - Same vehicle + same dealer -> same suffix
    - Same vehicle + new dealer -> next suffix
    """
    cn_numbers = []
    prev_vehicle = None
    dealer_suffix_map = {}

    for _, row in df.iterrows():
        vehicle = clean_text(row["Plate no"]).upper()
        dealer = clean_text(row["Distributor Code"]).upper()
        base_cn = make_base_cn_no(row["Lot no"])

        if prev_vehicle is None or vehicle != prev_vehicle:
            cn_numbers.append(base_cn)
            prev_vehicle = vehicle
            dealer_suffix_map = {dealer: ""}
        else:
            if dealer in dealer_suffix_map:
                suffix = dealer_suffix_map[dealer]
                cn_numbers.append(f"{base_cn}{suffix}")
            else:
                suffix = f"-{len(dealer_suffix_map)}"
                dealer_suffix_map[dealer] = suffix
                cn_numbers.append(f"{base_cn}{suffix}")

    return cn_numbers


def process_hyundai_cn(trpspn_file, filter_date=None):
    print("=" * 60)
    print("HYUNDAI CN PROCESS STARTED")
    print("=" * 60)
    print("TRPSPN file:", trpspn_file)

    trpspn_df = pd.read_excel(trpspn_file)
    trpspn_df = prepare_trpspn_df(trpspn_df)

    print("TRPSPN rows after cleanup:", len(trpspn_df))

    if filter_date:
        trpspn_df["parsed_gate_out"] = pd.to_datetime(
            trpspn_df["Gate Out Date"], errors="coerce", dayfirst=True
        )
        filter_dt = pd.to_datetime(filter_date, errors="coerce")
        if pd.notna(filter_dt):
            trpspn_df = trpspn_df[trpspn_df["parsed_gate_out"].dt.date == filter_dt.date()]

    if trpspn_df.empty:
        print("No data found after applying filter.")
        return None

    trpspn_df = trpspn_df.reset_index(drop=True)
    trpspn_df["FINAL_CN_NO"] = build_cn_numbers(trpspn_df)

    wb = load_workbook(TEMPLATE_FILE)
    ws = wb.active

    for row_idx in range(2, ws.max_row + 1):
        for col in range(1, 27):
            ws.cell(row_idx, col).value = None

    for i, (_, row) in enumerate(trpspn_df.iterrows(), start=2):
        plate_no = clean_text(row["Plate no"])
        lot_no = clean_text(row["Lot no"])
        cn_no = clean_text(row["FINAL_CN_NO"])
        cn_date = parse_date(row.get("Gate Out Date"))
        vin_no = clean_text(row.get("VIN No"))
        model_code = clean_text(row.get("Model"))
        mm_material = normalize_model_code(model_code)
        billing_date = parse_date(row.get("Billing date"))

        purch_amt = row.get("Purch Amt")
        rate = pd.to_numeric(purch_amt, errors="coerce")
        rate_value = "" if pd.isna(rate) else float(rate)
        freight_value = "" if pd.isna(rate) else float(rate)

        actual_route = f"{FIXED_OFFICE}-{FIXED_DESTINATION}-{FIXED_CONSIGNOR}"
        charged_route = actual_route

        print("-" * 60)
        print(f"Row {i}")
        print("Plate No:", plate_no)
        print("Lot No:", lot_no)
        print("CN No:", cn_no)
        print("Consignee:", FIXED_CONSIGNEE)
        print("Destination:", FIXED_DESTINATION)
        print("Billing Date:", billing_date)
        print("CN Date:", cn_date)
        print("VIN No:", vin_no)
        print("Model Code:", model_code)
        print("MM Material:", mm_material)
        print("Purch Amt:", purch_amt)
        
                    # Write to Excel
        ws.cell(i, 1).value = cn_no
        ws.cell(i, 2).value = FIXED_CN_TYPE
        ws.cell(i, 3).value = cn_date
        ws.cell(i, 4).value = FIXED_OFFICE
        ws.cell(i, 5).value = FIXED_BILLING_OFFICE
        ws.cell(i, 6).value = FIXED_CONSIGNOR
        ws.cell(i, 7).value = actual_route
        ws.cell(i, 8).value = charged_route
        ws.cell(i, 9).value = FIXED_CONSIGNEE
        ws.cell(i, 10).value = plate_no
        ws.cell(i, 11).value = FIXED_RATE_CHART
        ws.cell(i, 12).value = FIXED_LOAD_TYPE
        ws.cell(i, 13).value = ""
        ws.cell(i, 14).value = lot_no
        ws.cell(i, 15).value = billing_date
        ws.cell(i, 16).value = mm_material
        ws.cell(i, 17).value = FIXED_WEIGHT
        ws.cell(i, 18).value = rate_value
        ws.cell(i, 19).value = freight_value
        ws.cell(i, 20).value = ""
        ws.cell(i, 21).value = vin_no
        ws.cell(i, 22).value = ""
        ws.cell(i, 23).value = ""
        ws.cell(i, 24).value = ""
        ws.cell(i, 25).value = ""
        ws.cell(i, 26).value = ""
    
        # ===== NEW: SAVE TO MYSQL =====
        row_dict = {
            'CN No': cn_no,
            'CN Type': FIXED_CN_TYPE,
            'CN Date': cn_date,
            'Office List': FIXED_OFFICE,
            'Billing Office List': FIXED_BILLING_OFFICE,
            'Consignor': FIXED_CONSIGNOR,
            'Actual Route': actual_route,
            'Charged Route': charged_route,
            'Consignee': FIXED_CONSIGNEE,
            'Vehicle No': plate_no,
            'Rate Chart': FIXED_RATE_CHART,
            'Load Type': FIXED_LOAD_TYPE,
            'Lr No': None,
            'MM Invoice No': lot_no,
            'MM Invoice Date': billing_date,
            'MM Material': mm_material,
            'MM Actual Weight': FIXED_WEIGHT,
            'Rate': rate_value,
            'Freight': freight_value,
            'Other Charges': None,
            'MM Chassis No': vin_no,
            'MM Engine No': None,
            'MM NVRR No': None,
            'MM Remark': None,
            'No of Vehicles In Trailer': None,
            'Pod Date': None,
        }
        
        save_cn_to_mysql(row_dict)
        ws.cell(i, 1).value = cn_no
        ws.cell(i, 2).value = FIXED_CN_TYPE
        ws.cell(i, 3).value = cn_date
        ws.cell(i, 4).value = FIXED_OFFICE
        ws.cell(i, 5).value = FIXED_BILLING_OFFICE
        ws.cell(i, 6).value = FIXED_CONSIGNOR
        ws.cell(i, 7).value = actual_route
        ws.cell(i, 8).value = charged_route
        ws.cell(i, 9).value = FIXED_CONSIGNEE
        ws.cell(i, 10).value = plate_no
        ws.cell(i, 11).value = FIXED_RATE_CHART
        ws.cell(i, 12).value = FIXED_LOAD_TYPE
        ws.cell(i, 13).value = ""
        ws.cell(i, 14).value = lot_no
        ws.cell(i, 15).value = billing_date
        ws.cell(i, 16).value = mm_material
        ws.cell(i, 17).value = FIXED_WEIGHT
        ws.cell(i, 18).value = rate_value
        ws.cell(i, 19).value = freight_value
        ws.cell(i, 20).value = ""
        ws.cell(i, 21).value = vin_no
        ws.cell(i, 22).value = ""
        ws.cell(i, 23).value = ""
        ws.cell(i, 24).value = ""
        ws.cell(i, 25).value = ""
        ws.cell(i, 26).value = ""

    run_datetime = datetime.now().strftime("%Y-%m-%d_%H%M%S")
    output_file = f"HUNDAI_EXPORT_CN_{run_datetime}.xlsx"

    wb.save(output_file)

    archive_old_daily_files()

    final_output_path = os.path.join(DAILY_REPORTS_DIR, output_file)
    shutil.move(output_file, final_output_path)

    print("=" * 60)
    print(f"Saved to DailyReports: {final_output_path}")
    print("=" * 60)

    return final_output_path


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python Hundai_cn_export.py <trpspn_file.xlsx> [YYYY-MM-DD]")
        sys.exit(1)

    trpspn_file = sys.argv[1]
    filter_date = sys.argv[2] if len(sys.argv) > 2 else None

    result = process_hyundai_cn(trpspn_file, filter_date=filter_date)
    print("Output file generated:", result)